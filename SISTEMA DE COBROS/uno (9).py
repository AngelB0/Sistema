# @autor: JORGE ANGEL BECERRIL GONZALEZ, NOE REDZ MONDRAGRON
from customtkinter import CTk, CTkFrame, CTkEntry, CTkLabel, CTkButton, CTkCheckBox, CTkComboBox
import customtkinter as ctk
from tkinter import *
from customtkinter import CTk, CTkComboBox
import customtkinter as ctk
import tkinter as tk
from tkinter import ttk
from datetime import date
from tkinter import PhotoImage
from tkinter import filedialog
import pandas as pd
from sqlalchemy import create_engine
import pathlib
import sys
from tkinter import messagebox #libreria para los mensajes de alerta
import psycopg2 #importacion del modulo
import time#para el tiempo
from PIL import Image, ImageTk

from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import fitz  # PyMuPDF

from openpyxl import Workbook#pip install openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import fitz  # PyMuPDF
from openpyxl import Workbook#pip install openpyxl
from tkinter import simpledialog


#CLASE MAIN todo Va dentro de esta CLASS
def INDEX():
    #VENTANA PARA ELIMINAR 
    def eliminar():
        def delete_Alum_Uno():
            def dele_Alum():
                user3 = e1.get()
                passw3 = e2.get()
                bd="sistema"
                conexion = psycopg2.connect(
                                    user=user3,
                                    password=passw3,
                                    host='localhost',
                                    port='5432',
                                    database= bd
                                )
                cursor = conexion.cursor()

                pk_valor = int(matri3.get())

                # Ejecutar la sentencia DELETE
                tabla = "alumnos"  # Reemplaza con el nombre de tu tabla
                consulta = f"DELETE FROM {tabla} WHERE matricula = %s;"  # Reemplaza 'id' con el nombre de tu columna de clave primaria
                cursor.execute(consulta, (pk_valor,))

                matri3.delete(0, tk.END)
                alumno_a.delete(1.0, tk.END)
                plantel_a.delete(1.0, tk.END)
                ofedu_a.delete(1.0, tk.END)
                grad_a.delete(1.0, tk.END)
                grup_a.delete(1.0, tk.END)
                perido_a.delete(1.0, tk.END)
                turno_al.delete(1.0, tk.END)
                inscripcion_a.delete(1.0, tk.END)
                estatus_a.delete(1.0, tk.END)
                festatus_a.delete(1.0, tk.END)

                conexion.commit()
                cursor.close()
                conexion.close()

            def obtener_registro_completo_del(llave_primaria):
                user3 = e1.get()
                passw3 = e2.get()
                bd="sistema"
                conexion = psycopg2.connect(
                                    user=user3,
                                    password=passw3,
                                    host='localhost',
                                    port='5432',
                                    database= bd
                                )

                cursor = conexion.cursor()

                cursor.execute("SELECT * FROM alumnos WHERE Matricula = %s", (llave_primaria,))
                registro_completo = cursor.fetchone()

                cursor.close()
                conexion.close()

                return registro_completo

            def accion_al_seleccionar_del(event):
            # Añade más vectores según sea necesario para las demás columnas
                opcion_seleccionada = matri3.get()
                # Actualizar el contenido del Label con la opción seleccionada
                #label_resultado.config(text=f"Opción seleccionada: {opcion_seleccionada}")
                if opcion_seleccionada:
                    registro_completo = obtener_registro_completo_del(opcion_seleccionada)
                    if registro_completo:
                        #segun la matricula imprimimos la informacion del alumno
                        entered_text_1 = str(registro_completo[1])#ALUMNO
                        alumno_a.config(state=tk.NORMAL)
                        alumno_a.delete(1.0, tk.END)
                        alumno_a.insert(tk.END, f"{entered_text_1}")

                        entered_text_2 = str(registro_completo[2])#PLANTEL
                        plantel_a.config(state=tk.NORMAL)
                        plantel_a.delete(1.0, tk.END)
                        plantel_a.insert(tk.END, f"{entered_text_2}")

                        entered_text_3 = str(registro_completo[3])#OFERTA EDUCATIVA
                        ofedu_a.config(state=tk.NORMAL)
                        ofedu_a.delete(1.0, tk.END)
                        ofedu_a.insert(tk.END, f"{entered_text_3}")

                        entered_text_4 = str(registro_completo[4])#GRADO
                        grad_a.config(state=tk.NORMAL)
                        grad_a.delete(1.0, tk.END)
                        grad_a.insert(tk.END, f"{entered_text_4}")

                        entered_text_5 = str(registro_completo[5])#GRUPO
                        grup_a.config(state=tk.NORMAL)
                        grup_a.delete(1.0, tk.END)
                        grup_a.insert(tk.END, f"{entered_text_5}")
                        entered_text_6 = str(registro_completo[6])#PERIODO
                        perido_a.config(state=tk.NORMAL)
                        perido_a.delete(1.0, tk.END)
                        perido_a.insert(tk.END, f"{entered_text_6}")
                        entered_text_7 = str(registro_completo[7])#TURNO
                        turno_al.config(state=tk.NORMAL)
                        turno_al.delete(1.0, tk.END)
                        turno_al.insert(tk.END, f"{entered_text_7}")
                        entered_text_8 = str(registro_completo[8])#INSCRIPCION
                        inscripcion_a.config(state=tk.NORMAL)
                        inscripcion_a.delete(1.0, tk.END)
                        inscripcion_a.insert(tk.END, f"{entered_text_8}")
                        entered_text_9 = str(registro_completo[9])#ESTATUS
                        estatus_a.config(state=tk.NORMAL)
                        estatus_a.delete(1.0, tk.END)
                        estatus_a.insert(tk.END, f"{entered_text_9}")
                        entered_text_10 = str(registro_completo[10])#TURNO
                        festatus_a.config(state=tk.NORMAL)
                        festatus_a.delete(1.0, tk.END)
                        festatus_a.insert(tk.END, f"{entered_text_10}")
                    else:
                        print("Registro no encontrado.")

            def obtener_mat_bd_del():
                user3 = e1.get()
                passw3 = e2.get()
                bd="sistema"
                conexion = psycopg2.connect(
                                    user=user3,
                                    password=passw3,
                                    host='localhost',
                                    port='5432',
                                    database= bd
                                )
                # Crear un cursor para ejecutar consultas SQL
                cursor = conexion.cursor()

                # Ejecutar una consulta para obtener opciones (reemplaza con tu propia consulta)
                cursor.execute("SELECT * FROM alumnos")
                matris = cursor.fetchall()

                # Cerrar la conexión y el cursor
                cursor.close()
                conexion.close()

                return matris

            def on_key_release_del(event):
                    number = mostrar_mat_del()
                    try:
                        value = int(combobox_var.get())
                        filtered_values = [item for item in number if str(value) in str(item)]

                    except ValueError:
                        filtered_values = []

                    matri3['values'] = filtered_values

            def mostrar_mat_del():
                # Obtener las opciones de la base de datos
                opciones_bd_mat = obtener_mat_bd_del()
                # Limpiar el ComboBox
                matri3['values'] = ()
                # Configurar las nuevas opciones en el ComboBox
                matri3['values'] = [opcion[0] for opcion in opciones_bd_mat]
                number_list = matri3['values'] = [opcion[0] for opcion in opciones_bd_mat]

                return number_list
            
            # Crear una ventana secundaria.
            window_alu_del = tk.Toplevel()
            #color de fondo BG
            window_alu_del.config(bg=color_defondo)
            #se le añade un titulo a la app
            window_alu_del.title("ELIMINAR REGISTRO ALUMNOS")
            window_alu_del.geometry("1080x600+150+60")

            #MATRICULA
            tk.Label(window_alu_del, text="MATRICULA:", bg="#1D1212", fg="white").place(x=44, y=70, width=80, height=20)
            combobox_var = tk.StringVar()
            matri3 = ttk.Combobox(window_alu_del, width=20, textvariable=combobox_var)
            matri3.place(x=126, y=70, width=92, height=20)
            # Vincular la función al evento <<ComboboxSelected>>
            matri3.bind("<<ComboboxSelected>>", accion_al_seleccionar_del)
            matri3.bind('<KeyRelease>', on_key_release_del)

            #ALUMNO
            tk.Label(window_alu_del, text="ALUMNO:",bg="#1D1212", fg="white").place(x=36, y=100, width=80, height=20)
            alumno_a = tk.Text(window_alu_del, height=1, width=35, wrap=tk.NONE, bd=0, padx=2, pady=2)
            alumno_a.place(x=125, y=100)
            alumno_a.insert(tk.END, "")

            #OFERTA EDUCATIVA 
            tk.Label(window_alu_del, text="OFERTA EDUCATIVA:", bg="#1D1212", fg="white").place(x=420, y=100, width=120, height=20)
            ofedu_a = tk.Text(window_alu_del, height=1, width=50, wrap=tk.NONE, bd=0, padx=2, pady=2)
            ofedu_a.place(x=545, y=100)
            ofedu_a.insert(tk.END, "")

            #PLANTEL
            tk.Label(window_alu_del, text="PLANTEL:", bg="#1D1212", fg="white").place(x=690, y=130, width=75, height=20)
            plantel_a = tk.Text(window_alu_del, height=1, width=22, wrap=tk.NONE, bd=0, padx=2, pady=2)
            plantel_a.insert(tk.END, "")
            plantel_a.place(x=770, y=130)

            #GRADO
            tk.Label(window_alu_del, text="GRADO:",bg="#1D1212", fg="white").place(x=50, y=130, width=40, height=20)
            grad_a = tk.Text(window_alu_del, height=1, width=2, wrap=tk.NONE, bd=0, padx=2, pady=2)
            grad_a.insert(tk.END, "")
            grad_a.place(x=100, y=130)

            #GRUPO
            tk.Label(window_alu_del, text="GRUPO:", bg="#1D1212", fg="white").place(x=135, y=130, width=40, height=20)
            grup_a = tk.Text(window_alu_del, height=1, width=16, wrap=tk.NONE, bd=0, padx=2, pady=2)
            grup_a.insert(tk.END, "")
            grup_a.place(x=185, y=130)

            #PERIODO
            tk.Label(window_alu_del, text="PERIODO:",bg="#1D1212", fg="white").place(x=335, y=130, width=50, height=20)
            perido_a = tk.Text(window_alu_del, height=1, width=15, wrap=tk.NONE, bd=0, padx=2, pady=2)
            perido_a.insert(tk.END, "")
            perido_a.place(x=395, y=130)

            #TURNO
            tk.Label(window_alu_del, text="TURNO:",bg="#1D1212", fg="white").place(x=530, y=130, width=50, height=20)
            turno_al = tk.Text(window_alu_del, height=1, width=10, wrap=tk.NONE, bd=0, padx=2, pady=2)
            turno_al.insert(tk.END, "")
            turno_al.place(x=590, y=130)

            #FECHA DE INSCRIPCION
            tk.Label(window_alu_del, text="FECHA DE INSCRIPCION:", bg="#1D1212", fg="white").place(x=50, y=160, width=133, height=20)
            inscripcion_a = tk.Text(window_alu_del, height=1, width=10, wrap=tk.NONE, bd=0, padx=2, pady=2)
            inscripcion_a.insert(tk.END, "")
            inscripcion_a.place(x=190, y=160)

            #ESTATUS
            tk.Label(window_alu_del, text="ESTATUS:", bg="#1D1212", fg="white").place(x=285, y=160, width=80, height=20)
            estatus_a = tk.Text(window_alu_del, height=1, width=11, wrap=tk.NONE, bd=0, padx=2, pady=2)
            estatus_a.insert(tk.END, "")
            estatus_a.place(x=362, y=160)

            #FECHA ESTATUS
            tk.Label(window_alu_del, text="FECHA ESTATUS:", bg="#1D1212", fg="white").place(x=475, y=160, width=90, height=20)
            festatus_a = tk.Text(window_alu_del, height=1, width=10, wrap=tk.NONE, bd=0, padx=2, pady=2)
            festatus_a.insert(tk.END, "")
            festatus_a.place(x=570, y=160)

            #Boton cerrar
            boton_cerrar = tk.Button(window_alu_del, text="Cerrar ventana",bg="#0020BE", fg="white", relief=tk.RAISED, bd=5, command=window_alu_del.destroy)
            boton_cerrar.place(x=50, y=300, width=90, height=40)

            #Boton guardar
            boton_guardar = tk.Button(window_alu_del, text="ELIMINAR", bg="red", fg="white", relief=tk.RAISED, bd=5, command= dele_Alum)
            boton_guardar.place(x=150, y=300, width=90, height=40)

            # Cierra la ventana principal
            mostrar_mat_del()
            ventana.withdraw()
            
        def delete_Alum_com():
            try:
                # Establecer la conexión a la base de datos
                user3 = e1.get()
                passw3 = e2.get()
                bd="sistema"
                conn = psycopg2.connect(
                                    user=user3,
                                    password=passw3,
                                    host='localhost',
                                    port='5432',
                                    database= bd
                                )

                # Crear un objeto cursor
                cursor = conn.cursor()

                # Ejecutar la consulta para eliminar todos los registros de la tabla
                cursor.execute("DELETE FROM alumnos;")

                # Confirmar la transacción
                conn.commit()

                # Cerrar el cursor y la conexión
                cursor.close()
                conn.close()

                # Mostrar una alerta de éxito
                messagebox.showinfo("Éxito (Alumnos)", "Tabla eliminada correctamente.")

            except Exception as e:
                # Mostrar una alerta de error
                messagebox.showerror("Error", f"Error al eliminar registros:\n{str(e)}")

        #VENTANA ACTUALIZAR REGISTRO CONCEPTO
        def delete_Concep_Uno():
            def del_Concep():
                user3 = e1.get()
                passw3 = e2.get()
                bd="sistema"
                conexion = psycopg2.connect(
                                    user=user3,
                                    password=passw3,
                                    host='localhost',
                                    port='5432',
                                    database= bd
                                )
                cursor = conexion.cursor()
        
                pk_valor = int(codigo_a.get())

                # Ejecutar la sentencia DELETE
                tabla = "conceptos"  # Reemplaza con el nombre de tu tabla
                consulta1 = f"DELETE FROM {tabla} WHERE codigo = %s;"  # Reemplaza 'id' con el nombre de tu columna de clave primaria
                cursor.execute(consulta1, (pk_valor,))

                messagebox.showinfo("VALIDACION", "DATOS ELIMINADOS")
                
                codigo_a.delete(0, tk.END)
                concepto_a.delete(1.0, tk.END)
                contable_a.delete(1.0, tk.END)
                costo_a.delete(1.0, tk.END)
                clasificacion_a.delete(1.0, tk.END)
                conexion.commit()
                cursor.close()
                conexion.close()

            def obtener_registro_completo_con(llave_primaria):
                user3 = e1.get()
                passw3 = e2.get()
                bd="sistema"
                conexion = psycopg2.connect(
                                    user=user3,
                                    password=passw3,
                                    host='localhost',
                                    port='5432',
                                    database= bd
                                )

                cursor = conexion.cursor()

                cursor.execute("SELECT * FROM conceptos WHERE codigo = %s", (llave_primaria,))
                registro_completo = cursor.fetchone()

                cursor.close()
                conexion.close()

                return registro_completo

            def accion_al_seleccionar_con_del(event):
            # Añade más vectores según sea necesario para las demás columnas
                opcion_seleccionada = codigo_a.get()
                # Actualizar el contenido del Label con la opción seleccionada
                #label_resultado.config(text=f"Opción seleccionada: {opcion_seleccionada}")
                if opcion_seleccionada:
                    registro_completo = obtener_registro_completo_con(opcion_seleccionada)
                    if registro_completo:
                        #segun la matricula imprimimos la informacion del alumno
                        entered_text_1 = str(registro_completo[1])#CONCEPTO
                        concepto_a.config(state=tk.NORMAL)
                        concepto_a.delete(1.0, tk.END)
                        concepto_a.insert(tk.END, f"{entered_text_1}")

                        entered_text_2 = str(registro_completo[2])#CUENTA CONTABLE
                        contable_a.config(state=tk.NORMAL)
                        contable_a.delete(1.0, tk.END)
                        contable_a.insert(tk.END, f"{entered_text_2}")

                        entered_text_3 = str(registro_completo[3])#COSTO
                        costo_a.config(state=tk.NORMAL)
                        costo_a.delete(1.0, tk.END)
                        costo_a.insert(tk.END, f"{entered_text_3}")

                        entered_text_4 = str(registro_completo[4])#CLASIFICACION
                        clasificacion_a.config(state=tk.NORMAL)
                        clasificacion_a.delete(1.0, tk.END)
                        clasificacion_a.insert(tk.END, f"{entered_text_4}")
                    else:
                        print("Registro no encontrado.")

            def obtener_mat_bd_con_del():
                user3 = e1.get()
                passw3 = e2.get()
                bd="sistema"
                conexion = psycopg2.connect(
                                    user=user3,
                                    password=passw3,
                                    host='localhost',
                                    port='5432',
                                    database= bd
                                )
                # Crear un cursor para ejecutar consultas SQL
                cursor = conexion.cursor()

                # Ejecutar una consulta para obtener opciones (reemplaza con tu propia consulta)
                cursor.execute("SELECT * FROM conceptos")
                matris = cursor.fetchall()

                # Cerrar la conexión y el cursor
                cursor.close()
                conexion.close()

                return matris

            def on_key_release_con_del(event):
                    number = mostrar_mat_con_del()
                    try:
                        value = int(combobox_con.get())
                        filtered_values = [item for item in number if str(value) in str(item)]

                    except ValueError:
                        filtered_values = []

                    codigo_a['values'] = filtered_values

            def mostrar_mat_con_del():
                # Obtener las opciones de la base de datos
                opciones_bd_mat1 = obtener_mat_bd_con_del()
                # Limpiar el ComboBox
                codigo_a['values'] = ()
                # Configurar las nuevas opciones en el ComboBox
                codigo_a['values'] = [opcion[0] for opcion in opciones_bd_mat1]
                number_list = codigo_a['values'] = [opcion[0] for opcion in opciones_bd_mat1]

                return number_list

            # Crear una ventana secundaria.
            window_con_Del = tk.Toplevel()
            #color de fondo BG
            window_con_Del.config(bg=color_defondo)
            #se le añade un titulo a la app
            window_con_Del.title("ELIMINAR REGISTRO CONCEPTOS")
            window_con_Del.geometry("1080x600+150+60")

            #CODIGO
            combobox_con = tk.StringVar()
            tk.Label(window_con_Del, text="CODIGO:", bg="#1D1212", fg="white").place(x=40, y=70, width=80, height=20)
            codigo_a = ttk.Combobox(window_con_Del, width=20, textvariable=combobox_con)
            #codigo_a = ctk.ComboBox(window_con, width=50)
            codigo_a.place(x=135, y=70)
            # Vincular la función al evento <<ComboboxSelected>>
            codigo_a.bind("<<ComboboxSelected>>", accion_al_seleccionar_con_del)
            codigo_a.bind('<KeyRelease>', on_key_release_con_del)

            #CONCEPTO
            tk.Label(window_con_Del, text="CONCEPTO:", bg="#1D1212", fg="white").place(x=50, y=100, width=80, height=20)
            concepto_a = tk.Text(window_con_Del, height=1, width=25, wrap=tk.NONE, bd=0, padx=2, pady=2)
            concepto_a.place(x=135, y=100)
            concepto_a.insert(tk.END, "")

            #CUENTA CONTABLE 
            tk.Label(window_con_Del, text="CUENTA CONTABLE:", bg="#1D1212", fg="white").place(x=350, y=100, width=120, height=20)
            contable_a = tk.Text(window_con_Del, height=1, width=3, wrap=tk.NONE, bd=0, padx=2, pady=2)
            contable_a.place(x=475, y=100)
            contable_a.insert(tk.END,"")

            #COSTO
            tk.Label(window_con_Del, text="COSTO:", bg="#1D1212", fg="white").place(x=520, y=100, width=75, height=20)
            costo_a = tk.Text(window_con_Del, height=1, width=8, wrap=tk.NONE, bd=0, padx=2, pady=2)
            costo_a.place(x=585, y=100)
            costo_a.insert(tk.END, "")

            #CLASIFICACION
            tk.Label(window_con_Del, text="CLASIFICACION:", bg="#1D1212", fg="white").place(x=50, y=130, width=100, height=20)
            clasificacion_a = tk.Text(window_con_Del, height=1, width=15, wrap=tk.NONE, bd=0, padx=2, pady=2)
            clasificacion_a.place(x=160, y=130)
            clasificacion_a.insert(tk.END, "")

            #Boton cerrar
            boton_cerrar = tk.Button( window_con_Del, text="Cerrar ventana",bg="#0020BE", fg="white", relief=tk.RAISED, bd=5, command=window_con_Del.destroy)
            boton_cerrar.place(x=50, y=200, width=90, height=40)
            #Boton Guardar
            boton_guardar = tk.Button(window_con_Del, text="ELIMINAR", bg="RED", fg="white", relief=tk.RAISED, bd=5, command=del_Concep)
            boton_guardar.place(x=150, y=200, width=90, height=40)
            # Cierra la ventana principal
            mostrar_mat_con_del()
            ventana.withdraw()

        def delete_Concep_com():
             try:
                # Establecer la conexión a la base de datos
                user3 = e1.get()
                passw3 = e2.get()
                bd="sistema"
                conn = psycopg2.connect(
                                    user=user3,
                                    password=passw3,
                                    host='localhost',
                                    port='5432',
                                    database= bd
                                )

                # Crear un objeto cursor
                cursor = conn.cursor()

                # Ejecutar la consulta para eliminar todos los registros de la tabla
                cursor.execute("DELETE FROM conceptos;")

                # Confirmar la transacción
                conn.commit()

                # Cerrar el cursor y la conexión
                cursor.close()
                conn.close()

                # Mostrar una alerta de éxito
                messagebox.showinfo("Éxito (Conceptos)", "Tabla eliminada correctamente.")

             except Exception as e:
                # Mostrar una alerta de error
                messagebox.showerror("Error", f"Error al eliminar registros:\n{str(e)}")

        # Crear una ventana secundaria.
        window_delete = tk.Toplevel()
        #color de fondo BG
        window_delete.config(bg=color_defondo)
        #se le añade un titulo a la app
        window_delete.title("ELIMINAR REGISTROS")
        #tamaño de la ventana
        window_delete.geometry("500x300+500+150")
        #window_delete.config(width=500, height=300)
        
        tk.Label(window_delete, text="ELIJA UNA OPCION", bg="#1D1212", fg="white").place(x=30, y=30)

        # Crear un botón dentro de la ventana secundaria
         # para cerrar la misma.
        tk.Label(window_delete, text="ALUMNOS", bg="#1D1212", fg="WHITE").place(x=190, y=90)

        btn_Alumnos_del = tk.Button(window_delete,text="Eliminar tabla conmpleta", 
            bg="red", fg="white", relief=tk.RAISED, bd=5,command=delete_Alum_com
        )
        btn_Alumnos_del_1 = tk.Button(window_delete,text="Eliminar 1 Registro", 
            bg="orange", fg="white", relief=tk.RAISED, bd=5,command=delete_Alum_Uno
        )
        #FUNCION ELIMINAR CONCEPTOS
        tk.Label(window_delete, text="CONCEPTOS", bg="#1D1212", fg="WHITE").place(x=350, y=90)
        btn_Conceptos_del = tk.Button(window_delete,text="Eliminar tabla completa", bg="red", fg="white", relief=tk.RAISED, bd=5, 
            command=delete_Concep_com
        )
        btn_Conceptos_del_1 = tk.Button(window_delete,text="Eliminar 1 registro", bg="orange", fg="white", relief=tk.RAISED, bd=5, 
            command=delete_Concep_Uno
        )

        btn_cerrar_del = tk.Button(window_delete,text="Cerrar ventana", bg="#0020BE", fg="white", relief=tk.RAISED, bd=5, 
            command=window_delete.destroy
        )
        
        btn_cerrar_del.place(x=50, y=120, width=90, height=40)
        btn_Alumnos_del.place(x=150, y=120, width=145, height=40)
        btn_Alumnos_del_1.place(x=150, y=170, width=145, height=40)
        btn_Conceptos_del.place(x=310, y=120, width=150, height=40)
        btn_Conceptos_del_1.place(x=310, y=170, width=150, height=40)
        # Cierra la ventana principal
        ventana.withdraw()

    #VENTANA ACTUALIZAR
    def actualizar():
        #VENTANA ACTUALIZAR REGISTRO DE ALUMNO
        def act_Alum():
            def Save_Alum():
                user3 = e1.get()
                passw3 = e2.get()
                bd="sistema"
                conexion = psycopg2.connect(
                                    user=user3,
                                    password=passw3,
                                    host='localhost',
                                    port='5432',
                                    database= bd
                                )
                cursor = conexion.cursor()

                aux1 = int(matri2.get())
                aux2 = alumno_a.get(1.0, tk.END)
                aux3 = plantel_a.get(1.0, tk.END)
                aux4 = ofedu_a.get(1.0, tk.END)
                aux5= int(grad_a.get(1.0, tk.END))
                aux6 = grup_a.get(1.0, tk.END)
                aux7 = perido_a.get(1.0, tk.END)
                aux8 = turno_al.get(1.0, tk.END)
                aux9 = inscripcion_a.get(1.0, tk.END)
                aux10 = estatus_a.get(1.0, tk.END)
                aux11 = festatus_a.get(1.0, tk.END)
                cursor.execute("UPDATE alumnos SET alumno = %s, plantel = %s, oferta_edu = %s, grado = %s, grupo = %s, periodo = %s, turno = %s, fecha_ins = %s , estatus = %s, fecha_est = %s WHERE matricula = %s",
                                (aux2, aux3, aux4, aux5, aux6, aux7, aux8, aux9, aux10, aux11, aux1))
                messagebox.showinfo("VALIDACION", "DATOS ACTUALIZADOS")
                matri2.delete(0, tk.END)
                alumno_a.delete(1.0, tk.END)
                plantel_a.delete(1.0, tk.END)
                ofedu_a.delete(1.0, tk.END)
                grad_a.delete(1.0, tk.END)
                grup_a.delete(1.0, tk.END)
                perido_a.delete(1.0, tk.END)
                turno_al.delete(1.0, tk.END)
                inscripcion_a.delete(1.0, tk.END)
                estatus_a.delete(1.0, tk.END)
                festatus_a.delete(1.0, tk.END)

                conexion.commit()
                cursor.close()
                conexion.close()

            def obtener_registro_completo_act(llave_primaria):
                user3 = e1.get()
                passw3 = e2.get()
                bd="sistema"
                conexion = psycopg2.connect(
                                    user=user3,
                                    password=passw3,
                                    host='localhost',
                                    port='5432',
                                    database= bd
                                )

                cursor = conexion.cursor()

                cursor.execute("SELECT * FROM alumnos WHERE Matricula = %s", (llave_primaria,))
                registro_completo = cursor.fetchone()

                cursor.close()
                conexion.close()

                return registro_completo

            def accion_al_seleccionar_act(event):
            # Añade más vectores según sea necesario para las demás columnas
                opcion_seleccionada = matri2.get()
                # Actualizar el contenido del Label con la opción seleccionada
                #label_resultado.config(text=f"Opción seleccionada: {opcion_seleccionada}")
                if opcion_seleccionada:
                    registro_completo = obtener_registro_completo_act(opcion_seleccionada)
                    if registro_completo:
                        #segun la matricula imprimimos la informacion del alumno
                        entered_text_1 = str(registro_completo[1])#ALUMNO
                        alumno_a.config(state=tk.NORMAL)
                        alumno_a.delete(1.0, tk.END)
                        alumno_a.insert(tk.END, f"{entered_text_1}")

                        entered_text_2 = str(registro_completo[2])#PLANTEL
                        plantel_a.config(state=tk.NORMAL)
                        plantel_a.delete(1.0, tk.END)
                        plantel_a.insert(tk.END, f"{entered_text_2}")

                        entered_text_3 = str(registro_completo[3])#OFERTA EDUCATIVA
                        ofedu_a.config(state=tk.NORMAL)
                        ofedu_a.delete(1.0, tk.END)
                        ofedu_a.insert(tk.END, f"{entered_text_3}")

                        entered_text_4 = str(registro_completo[4])#GRADO
                        grad_a.config(state=tk.NORMAL)
                        grad_a.delete(1.0, tk.END)
                        grad_a.insert(tk.END, f"{entered_text_4}")

                        entered_text_5 = str(registro_completo[5])#GRUPO
                        grup_a.config(state=tk.NORMAL)
                        grup_a.delete(1.0, tk.END)
                        grup_a.insert(tk.END, f"{entered_text_5}")
                        entered_text_6 = str(registro_completo[6])#PERIODO
                        perido_a.config(state=tk.NORMAL)
                        perido_a.delete(1.0, tk.END)
                        perido_a.insert(tk.END, f"{entered_text_6}")
                        entered_text_7 = str(registro_completo[7])#TURNO
                        turno_al.config(state=tk.NORMAL)
                        turno_al.delete(1.0, tk.END)
                        turno_al.insert(tk.END, f"{entered_text_7}")
                        entered_text_8 = str(registro_completo[8])#INSCRIPCION
                        inscripcion_a.config(state=tk.NORMAL)
                        inscripcion_a.delete(1.0, tk.END)
                        inscripcion_a.insert(tk.END, f"{entered_text_8}")
                        entered_text_9 = str(registro_completo[9])#ESTATUS
                        estatus_a.config(state=tk.NORMAL)
                        estatus_a.delete(1.0, tk.END)
                        estatus_a.insert(tk.END, f"{entered_text_9}")
                        entered_text_10 = str(registro_completo[10])#TURNO
                        festatus_a.config(state=tk.NORMAL)
                        festatus_a.delete(1.0, tk.END)
                        festatus_a.insert(tk.END, f"{entered_text_10}")
                    else:
                        print("Registro no encontrado.")

            def obtener_mat_bd_act():
                user3 = e1.get()
                passw3 = e2.get()
                bd="sistema"
                conexion = psycopg2.connect(
                                    user=user3,
                                    password=passw3,
                                    host='localhost',
                                    port='5432',
                                    database= bd
                                )
                # Crear un cursor para ejecutar consultas SQL
                cursor = conexion.cursor()

                # Ejecutar una consulta para obtener opciones (reemplaza con tu propia consulta)
                cursor.execute("SELECT * FROM alumnos")
                matris = cursor.fetchall()

                # Cerrar la conexión y el cursor
                cursor.close()
                conexion.close()

                return matris

            def on_key_release_act(event):
                    number = mostrar_mat_act()
                    try:
                        value = int(combobox_var.get())
                        filtered_values = [item for item in number if str(value) in str(item)]

                    except ValueError:
                        filtered_values = []

                    matri2['values'] = filtered_values

            def mostrar_mat_act():
                # Obtener las opciones de la base de datos
                opciones_bd_mat = obtener_mat_bd_act()
                # Limpiar el ComboBox
                matri2['values'] = ()
                # Configurar las nuevas opciones en el ComboBox
                matri2['values'] = [opcion[0] for opcion in opciones_bd_mat]
                number_list = matri2['values'] = [opcion[0] for opcion in opciones_bd_mat]

                return number_list
            
            # Crear una ventana secundaria.
            window_alu = tk.Toplevel()
            #color de fondo BG
            window_alu.config(bg=color_defondo)
            #se le añade un titulo a la app
            window_alu.title("ACTUALIZAR REGISTRO ALUMNOS")
            window_alu.geometry("1080x600+150+60")

            #MATRICULA
            tk.Label(window_alu, text="MATRICULA:", bg="#1D1212", fg="white").place(x=44, y=70, width=80, height=20)
            combobox_var = tk.StringVar()
            matri2 = ttk.Combobox(window_alu, width=20, textvariable=combobox_var)
            matri2.place(x=126, y=70, width=92, height=20)
            # Vincular la función al evento <<ComboboxSelected>>
            matri2.bind("<<ComboboxSelected>>", accion_al_seleccionar_act)
            matri2.bind('<KeyRelease>', on_key_release_act)

            #ALUMNO
            tk.Label(window_alu, text="ALUMNO:",bg="#1D1212", fg="white").place(x=36, y=100, width=80, height=20)
            alumno_a = tk.Text(window_alu, height=1, width=35, wrap=tk.NONE, bd=0, padx=2, pady=2)
            alumno_a.place(x=125, y=100)
            alumno_a.insert(tk.END, "")

            #OFERTA EDUCATIVA 
            tk.Label(window_alu, text="OFERTA EDUCATIVA:", bg="#1D1212", fg="white").place(x=420, y=100, width=120, height=20)
            ofedu_a = tk.Text(window_alu, height=1, width=50, wrap=tk.NONE, bd=0, padx=2, pady=2)
            ofedu_a.place(x=545, y=100)
            ofedu_a.insert(tk.END, "")

            #PLANTEL
            tk.Label(window_alu, text="PLANTEL:", bg="#1D1212", fg="white").place(x=690, y=130, width=75, height=20)
            plantel_a = tk.Text(window_alu, height=1, width=22, wrap=tk.NONE, bd=0, padx=2, pady=2)
            plantel_a.insert(tk.END, "")
            plantel_a.place(x=770, y=130)

            #GRADO
            tk.Label(window_alu, text="GRADO:",bg="#1D1212", fg="white").place(x=50, y=130, width=40, height=20)
            grad_a = tk.Text(window_alu, height=1, width=2, wrap=tk.NONE, bd=0, padx=2, pady=2)
            grad_a.insert(tk.END, "")
            grad_a.place(x=100, y=130)

            #GRUPO
            tk.Label(window_alu, text="GRUPO:", bg="#1D1212", fg="white").place(x=135, y=130, width=40, height=20)
            grup_a = tk.Text(window_alu, height=1, width=16, wrap=tk.NONE, bd=0, padx=2, pady=2)
            grup_a.insert(tk.END, "")
            grup_a.place(x=185, y=130)

            #PERIODO
            tk.Label(window_alu, text="PERIODO:",bg="#1D1212", fg="white").place(x=335, y=130, width=50, height=20)
            perido_a = tk.Text(window_alu, height=1, width=15, wrap=tk.NONE, bd=0, padx=2, pady=2)
            perido_a.insert(tk.END, "")
            perido_a.place(x=395, y=130)

            #TURNO
            tk.Label(window_alu, text="TURNO:",bg="#1D1212", fg="white").place(x=530, y=130, width=50, height=20)
            turno_al = tk.Text(window_alu, height=1, width=10, wrap=tk.NONE, bd=0, padx=2, pady=2)
            turno_al.insert(tk.END, "")
            turno_al.place(x=590, y=130)

            #FECHA DE INSCRIPCION
            tk.Label(window_alu, text="FECHA DE INSCRIPCION:", bg="#1D1212", fg="white").place(x=50, y=160, width=133, height=20)
            inscripcion_a = tk.Text(window_alu, height=1, width=10, wrap=tk.NONE, bd=0, padx=2, pady=2)
            inscripcion_a.insert(tk.END, "")
            inscripcion_a.place(x=190, y=160)

            #ESTATUS
            tk.Label(window_alu, text="ESTATUS:", bg="#1D1212", fg="white").place(x=285, y=160, width=80, height=20)
            estatus_a = tk.Text(window_alu, height=1, width=11, wrap=tk.NONE, bd=0, padx=2, pady=2)
            estatus_a.insert(tk.END, "")
            estatus_a.place(x=362, y=160)

            #FECHA ESTATUS
            tk.Label(window_alu, text="FECHA ESTATUS:", bg="#1D1212", fg="white").place(x=475, y=160, width=90, height=20)
            festatus_a = tk.Text(window_alu, height=1, width=10, wrap=tk.NONE, bd=0, padx=2, pady=2)
            festatus_a.insert(tk.END, "")
            festatus_a.place(x=570, y=160)

            #Boton cerrar
            boton_cerrar = tk.Button(window_alu, text="Cerrar ventana",bg="#0020BE", fg="white", relief=tk.RAISED, bd=5, command=window_alu.destroy)
            boton_cerrar.place(x=50, y=300, width=90, height=40)

            #Boton guardar
            boton_guardar = tk.Button(window_alu, text="GUARDAR", bg="#0020BE", fg="white", relief=tk.RAISED, bd=5, command= Save_Alum)
            boton_guardar.place(x=150, y=300, width=90, height=40)

            # Cierra la ventana principal
            mostrar_mat_act()
            ventana.withdraw()
        
        #VENTANA ACTUALIZAR REGISTRO CONCEPTO
        def act_Concep():
            def save_Concep():
                user3 = e1.get()
                passw3 = e2.get()
                bd="sistema"
                conexion = psycopg2.connect(
                                    user=user3,
                                    password=passw3,
                                    host='localhost',
                                    port='5432',
                                    database= bd
                                )
                cursor = conexion.cursor()

                a1 = int(codigo_a.get())
                a2 = concepto_a.get(1.0, tk.END)
                a3 = contable_a.get(1.0, tk.END)
                a4 = costo_a.get(1.0, tk.END)
                a5 = clasificacion_a.get(1.0, tk.END)
                cursor.execute("UPDATE conceptos SET concepto = %s, cuenta_cont = %s, costo = %s, clasif = %s WHERE codigo = %s",
                                (a2, a3, a4, a5, a1))
                messagebox.showinfo("VALIDACION", "DATOS ACTUALIZADOS")
                codigo_a.delete(0, tk.END)
                concepto_a.delete(1.0, tk.END)
                contable_a.delete(1.0, tk.END)
                costo_a.delete(1.0, tk.END)
                clasificacion_a.delete(1.0, tk.END)
                conexion.commit()
                cursor.close()
                conexion.close()

            def obtener_registro_completo_act(llave_primaria):
                user3 = e1.get()
                passw3 = e2.get()
                bd="sistema"
                conexion = psycopg2.connect(
                                    user=user3,
                                    password=passw3,
                                    host='localhost',
                                    port='5432',
                                    database= bd
                                )

                cursor = conexion.cursor()

                cursor.execute("SELECT * FROM conceptos WHERE codigo = %s", (llave_primaria,))
                registro_completo = cursor.fetchone()

                cursor.close()
                conexion.close()

                return registro_completo

            def accion_al_seleccionar_con(event):
            # Añade más vectores según sea necesario para las demás columnas
                opcion_seleccionada = codigo_a.get()
                # Actualizar el contenido del Label con la opción seleccionada
                #label_resultado.config(text=f"Opción seleccionada: {opcion_seleccionada}")
                if opcion_seleccionada:
                    registro_completo = obtener_registro_completo_act(opcion_seleccionada)
                    if registro_completo:
                        #segun la matricula imprimimos la informacion del alumno
                        entered_text_1 = str(registro_completo[1])#CONCEPTO
                        concepto_a.config(state=tk.NORMAL)
                        concepto_a.delete(1.0, tk.END)
                        concepto_a.insert(tk.END, f"{entered_text_1}")

                        entered_text_2 = str(registro_completo[2])#CUENTA CONTABLE
                        contable_a.config(state=tk.NORMAL)
                        contable_a.delete(1.0, tk.END)
                        contable_a.insert(tk.END, f"{entered_text_2}")

                        entered_text_3 = str(registro_completo[3])#COSTO
                        costo_a.config(state=tk.NORMAL)
                        costo_a.delete(1.0, tk.END)
                        costo_a.insert(tk.END, f"{entered_text_3}")

                        entered_text_4 = str(registro_completo[4])#CLASIFICACION
                        clasificacion_a.config(state=tk.NORMAL)
                        clasificacion_a.delete(1.0, tk.END)
                        clasificacion_a.insert(tk.END, f"{entered_text_4}")
                    else:
                        print("Registro no encontrado.")

            def obtener_mat_bd_con():
                user3 = e1.get()
                passw3 = e2.get()
                bd="sistema"
                conexion = psycopg2.connect(
                                    user=user3,
                                    password=passw3,
                                    host='localhost',
                                    port='5432',
                                    database= bd
                                )
                # Crear un cursor para ejecutar consultas SQL
                cursor = conexion.cursor()

                # Ejecutar una consulta para obtener opciones (reemplaza con tu propia consulta)
                cursor.execute("SELECT * FROM conceptos")
                matris = cursor.fetchall()

                # Cerrar la conexión y el cursor
                cursor.close()
                conexion.close()

                return matris

            def on_key_release_con(event):
                    number = mostrar_mat_con()
                    try:
                        value = int(combobox_con.get())
                        filtered_values = [item for item in number if str(value) in str(item)]

                    except ValueError:
                        filtered_values = []

                    codigo_a['values'] = filtered_values

            def mostrar_mat_con():
                # Obtener las opciones de la base de datos
                opciones_bd_mat1 = obtener_mat_bd_con()
                # Limpiar el ComboBox
                codigo_a['values'] = ()
                # Configurar las nuevas opciones en el ComboBox
                codigo_a['values'] = [opcion[0] for opcion in opciones_bd_mat1]
                number_list = codigo_a['values'] = [opcion[0] for opcion in opciones_bd_mat1]

                return number_list

            # Crear una ventana secundaria.
            window_con = tk.Toplevel()
            #color de fondo BG
            window_con.config(bg=color_defondo)
            #se le añade un titulo a la app
            window_con.title("ACTUALIZAR REGISTRO CONCEPTOS")
            window_con.geometry("1080x600+150+60")

            #CODIGO
            combobox_con = tk.StringVar()
            tk.Label(window_con, text="CODIGO:", bg="#1D1212", fg="white").place(x=40, y=70, width=80, height=20)
            codigo_a = ttk.Combobox(window_con, width=20, textvariable=combobox_con)
            #codigo_a = ctk.ComboBox(window_con, width=50)
            codigo_a.place(x=135, y=70)
            # Vincular la función al evento <<ComboboxSelected>>
            codigo_a.bind("<<ComboboxSelected>>", accion_al_seleccionar_con)
            codigo_a.bind('<KeyRelease>', on_key_release_con)

            #CONCEPTO
            tk.Label(window_con, text="CONCEPTO:", bg="#1D1212", fg="white").place(x=50, y=100, width=80, height=20)
            concepto_a = tk.Text(window_con, height=1, width=25, wrap=tk.NONE, bd=0, padx=2, pady=2)
            concepto_a.place(x=135, y=100)
            concepto_a.insert(tk.END, "")

            #CUENTA CONTABLE 
            tk.Label(window_con, text="CUENTA CONTABLE:", bg="#1D1212", fg="white").place(x=350, y=100, width=120, height=20)
            contable_a = tk.Text(window_con, height=1, width=3, wrap=tk.NONE, bd=0, padx=2, pady=2)
            contable_a.place(x=475, y=100)
            contable_a.insert(tk.END,"")

            #COSTO
            tk.Label(window_con, text="COSTO:", bg="#1D1212", fg="white").place(x=520, y=100, width=75, height=20)
            costo_a = tk.Text(window_con, height=1, width=8, wrap=tk.NONE, bd=0, padx=2, pady=2)
            costo_a.place(x=585, y=100)
            costo_a.insert(tk.END, "")

            #CLASIFICACION
            tk.Label(window_con, text="CLASIFICACION:", bg="#1D1212", fg="white").place(x=50, y=130, width=100, height=20)
            clasificacion_a = tk.Text(window_con, height=1, width=15, wrap=tk.NONE, bd=0, padx=2, pady=2)
            clasificacion_a.place(x=160, y=130)
            clasificacion_a.insert(tk.END, "")

            #Boton cerrar
            boton_cerrar = tk.Button( window_con, text="Cerrar ventana",bg="#0020BE", fg="white", relief=tk.RAISED, bd=5, command=window_con.destroy)
            boton_cerrar.place(x=50, y=200, width=90, height=40)
            #Boton Guardar
            boton_guardar = tk.Button(window_con, text="GUARDAR", bg="#0020BE", fg="white", relief=tk.RAISED, bd=5, command=save_Concep)
            boton_guardar.place(x=150, y=200, width=90, height=40)
            # Cierra la ventana principal
            mostrar_mat_con()
            ventana.withdraw()
            
        # Crear una ventana secundaria.
        window_act = tk.Toplevel()
        #color de fondo BG
        window_act.config(bg=color_defondo)
        #se le añade un titulo a la app
        window_act.title("ACTUALIZAR REGISTROS")
        window_act.geometry("500x300+400+150")
                
        # Crear un botón dentro de la ventana secundaria
         # para cerrar la misma.
        btn_Alumnos_act = tk.Button(
            window_act,
            text="Actualizar Alumnos", 
            bg="#0020BE", fg="white", relief=tk.RAISED, bd=5, 
            command=act_Alum
         )
        btn_Conceptos_act = tk.Button(
            window_act,
            text="Actualizar Conceptos", 
            bg="#0020BE", fg="white", relief=tk.RAISED, bd=5, 
            command=act_Concep
        )
        btn_venta_act = tk.Button(
            window_act,
            text="Actualizar Venta", 
            bg="#0020BE", fg="white", relief=tk.RAISED, bd=5, 
            command=act_Concep
        )
        btn_cerrar_act = tk.Button(
            window_act,
            text="Cerrar ventana", 
            bg="#0020BE", fg="white", relief=tk.RAISED, bd=5, 
            command=window_act.destroy
        )
        
        btn_cerrar_act.place(x=50, y=100, width=90, height=40)
        btn_Alumnos_act.place(x=150, y=100, width=145, height=40)
        btn_Conceptos_act.place(x=310, y=100, width=150, height=40)
        btn_venta_act.place(x=50, y=150, width=150, height=40)
        # Cierra la ventana principal
        ventana.withdraw()
    
   #VENTANA PARA CONSULTAR
    def Consultar():
        ventana_consul=tk.Tk()
        #color de fondo BG
        ventana_consul.config(bg=color_defondo)
        #se le añade un titulo a la app
        ventana_consul.title("CONSULTA DE REGISTROS")
        ventana_consul.state("zoomed")

        # Variable de control para rastrear el estado del ComboBox de exportación
        combo_exportar_activado = tk.BooleanVar()
        combo_exportar_activado.set(False)

        # Variable de control para rastrear el estado del botón de ventas
        #btn_ventas_activado = tk.BooleanVar()
        #btn_ventas_activado.set(False)

        # Variable de control para rastrear el estado del botón de usuario
        #btn_usuario_activado = tk.BooleanVar()
        #btn_usuario_activado.set(False)

        # Variable de control para rastrear el estado del botón de fecha
        #btn_fecha_activado = tk.BooleanVar()
        #btn_fecha_activado.set(False)

        def ajustar_visibilidad_exportar_excel():
            if combo_exportar_activado.get():
                treeview_frame.combo_exportar.pack(side=tk.BOTTOM, pady=5)
            else:
                treeview_frame.combo_exportar.pack_forget()


        #def ajustar_visibilidad_exportar_excel():
            #if btn_ventas_activado.get():
                #treeview_frame.btn_exportar_excel.pack(side=tk.BOTTOM, pady=5)
                #if btn_usuario_activado.get():
                    #treeview_frame.btn_exportar_usuario.pack(side=tk.BOTTOM, pady=5)
                #else:
                    #treeview_frame.btn_exportar_usuario.pack_forget()
                
                #if btn_fecha_activado.get():
                    #treeview_frame.btn_exportar_fecha.pack(side=tk.BOTTOM, pady=5)
                #else:
                    #treeview_frame.btn_exportar_fecha.pack_forget()
            #else:
                #treeview_frame.btn_exportar_excel.pack_forget()
                #treeview_frame.btn_exportar_usuario.pack_forget()
                #treeview_frame.btn_exportar_fecha.pack_forget()
   

        def Consu_Alum():
            # Mostrar información de archivos
            treeview_frame.mostrar_info_archivos()
            # Mostrar información de alumnos
            treeview_frame.mostrar_info_alumnos()
            # Actualizar el estado del botón de ventas
            combo_exportar_activado.set(False)
            #btn_ventas_activado.set(False)
            #btn_usuario_activado.set(False)
            #btn_fecha_activado.set(False)
            # Ajustar la visibilidad del botón de exportar
            ajustar_visibilidad_exportar_excel()

        def Consu_Concep():
            # Mostrar información de conceptos
            treeview_frame.mostrar_info_conceptos()
            # Actualizar el estado del botón de ventas
            combo_exportar_activado.set(False)
            #btn_ventas_activado.set(False)
            #btn_usuario_activado.set(False)
            #btn_fecha_activado.set(False)
            # Ajustar la visibilidad del botón de exportar
            ajustar_visibilidad_exportar_excel()
        
        def Consu_vent():
            treeview_frame.mostrar_info_ventass()
            #Mostrar infromacion de ventas
            treeview_frame.mostrar_info_ventass()
            #Mostrar infromacion de ventas
            # Actualizar el estado del botón de ventas
            combo_exportar_activado.set(True)
            #btn_ventas_activado.set(True)
            #btn_usuario_activado.set(True)
            #btn_fecha_activado.set(True)
            # Ajustar la visibilidad del botón de exportar
            ajustar_visibilidad_exportar_excel()
        

        class TreeviewFrame(ttk.Frame):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)

                self.hscrollbar = ttk.Scrollbar(self, orient=tk.HORIZONTAL)
                self.vscrollbar = ttk.Scrollbar(self, orient=tk.VERTICAL)
                self.treeview = ttk.Treeview(
                    self,
                    xscrollcommand=self.hscrollbar.set,
                    yscrollcommand=self.vscrollbar.set
                )

                self.hscrollbar.config(command=self.treeview.xview)
                self.hscrollbar.pack(side=tk.BOTTOM, fill=tk.X)
                self.vscrollbar.config(command=self.treeview.yview)
                self.vscrollbar.pack(side=tk.RIGHT, fill=tk.Y)

                self.treeview.place(x=0, y=0, width=1265, height=540)

                # Atributo para rastrear el estado del Treeview (minimizado o no)
                self.minimizado = False

                # Reemplazar la creación del botón por la creación del ComboBox
                self.combo_exportar = ctk.CTkComboBox(self, values=["Reporte de Ventas", "Reporte de Ventas por Usuario", "Reporte de Ventas por Fecha"], command=self.ejecutar_funcion_seleccionada,
                                                      state="readonly",
                                                      hover=True,
                                                      height=28,
                                                      width=250,
                                                      font=("Candara", 14, "bold"),
                                                      fg_color='white',
                                                      text_color='black',
                                                      dropdown_font=("Candara", 14, "bold"),#Tamaño y letra de las opciones 
                                                      corner_radius=50,
                                                      border_width=3,
                                                      border_color='#00B471',
                                                      button_color='#00B471',
                                                      button_hover_color='white',
                                                      dropdown_fg_color='white',
                                                      dropdown_hover_color='#00CC80',
                                                      dropdown_text_color='black',
                                                      )
                #self.combo_exportar = ttk.Combobox(self, values=["Descargar todas las Ventas", "Descargar solo por Usuario", "Descargar solo por Fecha"], foreground='green', width=25, font=("Candara", 14, "bold"), state="readonly",)
                # Establecer un valor predeterminado (título por defecto)
                self.combo_exportar.set("Reporte de Ingresos")
                #self.combo_exportar.bind("<<ComboboxSelected>>", self.ejecutar_funcion_seleccionada)


                
            def ejecutar_funcion_seleccionada(self, event):
                opcion_seleccionada = self.combo_exportar.get()
                if opcion_seleccionada == "Reporte de Ventas":
                    self.exportar_a_excel()
                elif opcion_seleccionada == "Reporte de Ventas por Usuario":
                    self.exportar_a_excel_con_filtro_usuario()
                elif opcion_seleccionada == "Reporte de Ventas por Fecha":
                    self.exportar_a_excel_con_filtro_fecha()


                # Botón para exportar a Excel
                #self.btn_exportar_excel = tk.Button(self, text="Dercargar todas las Ventas",bg="#00A055", fg="white", relief=tk.RAISED, bd=5, command=self.exportar_a_excel)
                # Configurar el botón para que esté inicialmente oculto
                #self.btn_exportar_excel.pack_forget()

                #self.btn_exportar_usuario = tk.Button(self, text="Descargar solo por Usuario",bg="#00A055", fg="white", relief=tk.RAISED, bd=5, command=self.exportar_a_excel_con_filtro_usuario)
                #self.btn_exportar_usuario.pack_forget()

                #self.btn_exportar_fecha = tk.Button(self, text="Descargar solo por Fecha",bg="#00A055", fg="white", relief=tk.RAISED, bd=5, command=self.exportar_a_excel_con_filtro_fecha)
                #self.btn_exportar_fecha.pack_forget()

            def exportar_a_excel_con_filtro_fecha(self):
                # Obtener la fecha para filtrar
                fecha = simpledialog.askstring("Filtrar por Fecha", "Ingrese la fecha (formato: año-mes-dia  ejemplo: 2024-02-01):")
                if fecha is not None:
                    # Filtrar los datos por fecha
                    datos_filtrados = [item for item in self.treeview.get_children() if self.treeview.item(item, "values")[12] == fecha]

                    if not datos_filtrados:
                        messagebox.showinfo("Información", f"No hay datos para la fecha {fecha}")
                        return

                    # Obtener la ruta de guardado desde un cuadro de diálogo
                    ruta_archivo = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Archivos de Excel", "*.xlsx")])

                    if ruta_archivo:
                        # Crear un nuevo libro de Excel
                        libro_excel = Workbook()
                        # Crear una hoja en el libro
                        hoja_excel = libro_excel.active

                        # Obtener las columnas del TreeView
                        columnas = self.treeview["columns"]
                        # Escribir encabezados de columna en la hoja de Excel
                        for col, columna in enumerate(columnas):
                            hoja_excel.cell(row=1, column=col + 1, value=columna)

                        # Obtener los datos filtrados del TreeView y escribirlos en la hoja de Excel
                        for i, item in enumerate(datos_filtrados):
                            valores = self.treeview.item(item, "values")
                            for col, valor in enumerate(valores):
                                hoja_excel.cell(row=i + 2, column=col + 1, value=valor)

                        try:
                            # Guardar el archivo Excel
                            libro_excel.save(ruta_archivo)
                            messagebox.showinfo("Éxito", f"Los datos se han exportado correctamente a {ruta_archivo}")
                        except Exception as e:
                            messagebox.showerror("Error", f"Error al guardar el archivo: {e}")
            
            def obtener_usuarios(self):
                # Obtener la lista de usuarios únicos
                usuarios_unicos = set()
                for item in self.treeview.get_children():
                    usuario = self.treeview.item(item, "values")[14]
                    usuarios_unicos.add(usuario)
    
                    return list(usuarios_unicos)

            def exportar_a_excel_con_filtro_usuario(self):
                # Obtener el nombre de usuario para filtrar
                nombre_usuario = simpledialog.askstring("Filtrar por Usuario", "Ingrese el nombre de usuario:")
                if nombre_usuario is not None:
                    # Filtrar los datos por usuario
                    datos_filtrados = [item for item in self.treeview.get_children() if self.treeview.item(item, "values")[14] == nombre_usuario]

                    if not datos_filtrados:
                        messagebox.showinfo("Información", f"No hay datos para el usuario {nombre_usuario}")
                        return

                    # Obtener la ruta de guardado desde un cuadro de diálogo
                    ruta_archivo = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Archivos de Excel", "*.xlsx")])
                    # Configurar el tamaño de la ventana de filedialog

                    if ruta_archivo:
                        # Crear un nuevo libro de Excel
                        libro_excel = Workbook()
                        # Crear una hoja en el libro
                        hoja_excel = libro_excel.active

                        # Obtener las columnas del TreeView
                        columnas = self.treeview["columns"]
                        # Escribir encabezados de columna en la hoja de Excel
                        for col, columna in enumerate(columnas):
                            hoja_excel.cell(row=1, column=col + 1, value=columna)

                        # Obtener los datos filtrados del TreeView y escribirlos en la hoja de Excel
                        for i, item in enumerate(datos_filtrados):
                            valores = self.treeview.item(item, "values")
                            for col, valor in enumerate(valores):
                                hoja_excel.cell(row=i + 2, column=col + 1, value=valor)

                        try:
                            # Guardar el archivo Excel
                            libro_excel.save(ruta_archivo)
                            messagebox.showinfo("Éxito", f"Los datos se han exportado correctamente a {ruta_archivo}")
                        except Exception as e:
                            messagebox.showerror("Error", f"Error al guardar el archivo: {e}")

            def exportar_a_excel(self):
                # Obtener la ruta de guardado desde un cuadro de diálogo
                ruta_archivo = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Archivos de Excel", "*.xlsx")])
                

                if ruta_archivo:
                    # Crear un nuevo libro de Excel
                    libro_excel = Workbook()
                    # Crear una hoja en el libro
                    hoja_excel = libro_excel.active

                    # Obtener las columnas del TreeView
                    columnas = self.treeview["columns"]
                    # Escribir encabezados de columna en la hoja de Excel
                    for col, columna in enumerate(columnas):
                        hoja_excel.cell(row=1, column=col+1, value=columna)

                    # Obtener los datos del TreeView y escribirlos en la hoja de Excel
                    for i, item in enumerate(self.treeview.get_children()):
                        valores = self.treeview.item(item, "values")
                        for col, valor in enumerate(valores):
                            hoja_excel.cell(row=i + 2, column=col + 1, value=valor)

                    try:
                        # Guardar el archivo Excel
                        libro_excel.save(ruta_archivo)
                        messagebox.showinfo("Éxito", f"Los datos se han exportado correctamente a {ruta_archivo}")
                    except Exception as e:
                        messagebox.showerror("Error", f"Error al guardar el archivo: {e}")

            def mostrar_info_archivos(self):
                for file in pathlib.Path(sys.executable).parent.iterdir():
                    self.treeview.insert(
                        "", tk.END, values=(file.name, file.stat().st_size))

            def mostrar_info_alumnos(self):
                def obtener_datosBD():
                    #conexion a la BD
                    user3 = e1.get()
                    passw3 = e2.get()
                    bd="sistema"
                    conexion = psycopg2.connect(
                                user=user3,
                                password=passw3,
                                host='localhost',
                                port='5432',
                                database= bd
                            )
                    cursor = conexion.cursor()
                    cursor.execute("SELECT * FROM alumnos")
                    datos = cursor.fetchall()
                    cursor.close()
                    conexion.close()
                    return datos

                for fila in self.treeview.get_children():
                    self.treeview.delete(fila)

                self.treeview.config(columns=("matricula", "alumno", "plantel", "oferta_edu", "grado", "grupo", "periodo", "turno", "fecha_ins", "estatus", "fecha_est"), show="headings")
                self.treeview.column('#0', width=0, stretch=tk.NO)
                self.treeview.column('matricula', anchor=CENTER, width=80,stretch=tk.NO)
                self.treeview.column('alumno', width=220,stretch=tk.NO)
                self.treeview.column('plantel', anchor=CENTER, width=  120,stretch=tk.NO)
                self.treeview.column('oferta_edu', width=350,stretch=tk.NO)
                self.treeview.column('grado', anchor=CENTER, width=60,stretch=tk.NO)
                self.treeview.column('grupo', anchor=CENTER, width=110,stretch=tk.NO)
                self.treeview.column('periodo', anchor=CENTER, width=100,stretch=tk.NO)
                self.treeview.column('turno', anchor=CENTER, width=100, stretch=tk.NO)
                self.treeview.column('fecha_ins', anchor=CENTER, width=120, stretch=tk.NO)
                self.treeview.column('estatus', anchor=CENTER, width=100, stretch=tk.NO)
                self.treeview.column('fecha_est', anchor=CENTER, width=120, stretch=tk.NO)
                self.treeview.heading('#0', text='', anchor=CENTER)
                self.treeview.heading("matricula", text="Matricula")
                self.treeview.heading("alumno", text="Alumno")
                self.treeview.heading("plantel", text="Plantel")
                self.treeview.heading("oferta_edu", text="Oferta Educativa")
                self.treeview.heading("grado", text="Grado")
                self.treeview.heading("grupo", text="Grupo")
                self.treeview.heading("periodo", text="Periodo")
                self.treeview.heading("turno", text="Turno")
                self.treeview.heading("fecha_ins", text="Fecha de Inscripción")
                self.treeview.heading("estatus", text="Estatus")
                self.treeview.heading("fecha_est", text="Fecha de Estatus")

                datos = obtener_datosBD()
                for dato in datos:
                    self.treeview.insert("", "end", values=dato)

            def mostrar_info_conceptos(self):
                def obtener_datos_conceptos():
                    #conexion a la BD
                    user3 = e1.get()
                    passw3 = e2.get()
                    bd="sistema"
                    conexion = psycopg2.connect(
                                user=user3,
                                password=passw3,
                                host='localhost',
                                port='5432',
                                database= bd
                            )
                    cursor = conexion.cursor()
                    cursor.execute("SELECT * FROM conceptos")
                    datos = cursor.fetchall()
                    cursor.close()
                    conexion.close()
                    return datos

                for fila in self.treeview.get_children():
                    self.treeview.delete(fila)

                self.treeview.config(columns=("codigo", "concepto", "cuenta_cont", "costo", "clasif"), show="headings")
                self.treeview.column('#0', width=0)
                self.treeview.column('codigo', anchor=CENTER, width=80)
                self.treeview.column('concepto', anchor=CENTER, width=80)
                self.treeview.column('cuenta_cont', anchor=CENTER, width=80)
                self.treeview.column('costo', anchor=CENTER, width=80)
                self.treeview.column('clasif', anchor=CENTER, width=80)
                self.treeview.heading('#0', text='', anchor=CENTER)
                self.treeview.heading("codigo", text="Código")
                self.treeview.heading("concepto", text="Concepto")
                self.treeview.heading("cuenta_cont", text="Cuenta Contable")
                self.treeview.heading("costo", text="Costo")
                self.treeview.heading("clasif", text="Clasificación")

                datos = obtener_datos_conceptos()
                for dato in datos:
                    self.treeview.insert("", "end", values=dato)
            
            def mostrar_info_ventass(self):
                def obtener_datosBD_ventas():
                    #conexion a la BD
                    user3 = e1.get()
                    passw3 = e2.get()
                    bd="sistema"
                    conexion = psycopg2.connect(
                                user=user3,
                                password=passw3,
                                host='localhost',
                                port='5432',
                                database= bd
                            )
                    cursor = conexion.cursor()
                    cursor.execute("SELECT * FROM ventas")
                    datos = cursor.fetchall()
                    cursor.close()
                    conexion.close()
                    return datos

                for fila in self.treeview.get_children():
                    self.treeview.delete(fila)

                self.treeview.config(columns=("folio","plantel", "matri", "nombre", "grupo", "licenciatura", "cant_recibe", "forma_p", "aprov", "cuenta_rec", "observasiones", "subtotal", "fecha", "concep","usuario"), show="headings")
                self.treeview.column('#0', width=0, stretch=tk.NO)
                self.treeview.column('folio', anchor=CENTER, width=100,stretch=tk.NO)
                self.treeview.column('plantel', anchor=CENTER, width=100,stretch=tk.NO)
                self.treeview.column('matri', anchor=CENTER, width=100,stretch=tk.NO)
                self.treeview.column('nombre', anchor=CENTER, width=100,stretch=tk.NO)
                self.treeview.column('grupo', anchor=CENTER, width=100,stretch=tk.NO)
                self.treeview.column('licenciatura', anchor=CENTER, width=100,stretch=tk.NO)
                self.treeview.column('cant_recibe', anchor=CENTER, width=100,stretch=tk.NO)
                self.treeview.column('forma_p', anchor=CENTER, width=100,stretch=tk.NO)
                self.treeview.column('aprov', anchor=CENTER, width=100,stretch=tk.NO)
                self.treeview.column('cuenta_rec', anchor=CENTER, width=100,stretch=tk.NO)
                self.treeview.column('observasiones', anchor=CENTER, width=100,stretch=tk.NO)
                self.treeview.column('subtotal', anchor=CENTER, width=100,stretch=tk.NO)
                self.treeview.column('fecha', anchor=CENTER, width=100,stretch=tk.NO)
                self.treeview.column('concep', anchor=CENTER, width=100,stretch=tk.NO)
                self.treeview.column('usuario', anchor=CENTER, width=100,stretch=tk.NO)
                self.treeview.heading('#0', text='', anchor=CENTER) 
                self.treeview.heading("folio", text="Folio")
                self.treeview.heading("plantel", text="Plantel")
                self.treeview.heading("matri", text="Matricula")
                self.treeview.heading("nombre", text="Nombre Alumno")
                self.treeview.heading("grupo", text="Grupo")
                self.treeview.heading("licenciatura", text="Oferta Educativa")
                self.treeview.heading("cant_recibe", text="Cantidad que Recibio")
                self.treeview.heading("forma_p", text="Forma de Pago")
                self.treeview.heading("aprov", text="Aprobación")
                self.treeview.heading("cuenta_rec", text="Cuenta Reseptora")
                self.treeview.heading("observasiones", text="Observasiones")
                self.treeview.heading("subtotal", text="Precio Total")
                self.treeview.heading("fecha", text="Fecha de Venta")
                self.treeview.heading("concep", text="Conceptos")
                self.treeview.heading("usuario", text="Usuario")

                datos = obtener_datosBD_ventas()
                for dato in datos:
                    self.treeview.insert("", "end", values=dato)

                self.combo_exportar.pack(side=tk.BOTTOM, anchor=tk.E, padx=30)


                # Hacer que aparezca el botón exportar Excel
                #self.btn_exportar_excel.pack(side=tk.RIGHT, pady=40, anchor=tk.E, padx=30)
                # Hacer que aparezcan los botones exportar por usuario y exportar por fecha
                #self.btn_exportar_usuario.pack(side=tk.RIGHT, pady=40, anchor=tk.E, padx=200)
                #self.btn_exportar_fecha.pack(side=tk.RIGHT, pady=40, anchor=tk.E, padx=400)


            def toggle_minimizar(self):
                # Cambiar el estado de minimizado y mostrar u ocultar el Treeview en consecuencia
                self.minimizado = not self.minimizado
                if self.minimizado:
                    self.treeview.pack_forget()
                else:
                    self.treeview.pack()

        
        # Crear una instancia de la clase TreeviewFrame
        treeview_frame = TreeviewFrame(ventana_consul)
        # Empaquetar el TreeviewFrame en la ventana principal
        treeview_frame.place(x=0, y=0, width=1280, height=590)
        #treeview_frame.pack(fill=tk.BOTH, expand=True)

        # Crear un botón dentro de la ventana secundaria
        # para cerrar la misma.        
        btn_ventass = tk.Button(ventana_consul, text="Consultar Ventas", bg="#0020BE", fg="white", relief=tk.RAISED, bd=5, command=Consu_vent)    
        btn_Alumnos = tk.Button(ventana_consul, text="Consultar Alumnos", bg="#0020BE", fg="white", relief=tk.RAISED, bd=5, command=Consu_Alum)
        btn_Conceptos = tk.Button(ventana_consul,text="Consultar Conceptos",bg="#0020BE", fg="white", relief=tk.RAISED, bd=5, command=Consu_Concep)
        boton_regresar = tk.Button(ventana_consul,text="REGRESAR", bg="#0020BE", fg="white", relief=tk.RAISED, bd=5, command=lambda:close_window(ventana_consul))
        # Botón para exportar a Excel en el área de la ventana de ventas
        #btn_exportar_excel = tk.Button(ventana_consul, text="Exportar a Excel", command=treeview_frame.exportar_a_excel)
        #btn_exportar_excel.place(x=540, y=600, width=120, height=40)

        boton_regresar.place(x=50, y=600, width=90, height=40)
        btn_Alumnos.place(x=150, y=600, width=120, height=40)
        btn_Conceptos.place(x=280, y=600, width=125, height=40)
        btn_ventass.place(x=410, y=600, width=120, height=40)

    def close_window(window):
        # Cierra la ventana principal
        #WIND.withdraw()
        window.destroy()
        WIND.deiconify()
        
    def Insertar():
        def import_data_con():
            user4= e1.get()
            passw4 = e2.get()
            table_n1='conceptos'
            try:
                # Abrir el cuadro de diálogo para seleccionar el archivo Excel o CSV
                file_path = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx;*.xls"), ("Archivos CSV", "*.csv")])

                # Leer el archivo con pandas
                if file_path.endswith(".csv"):
                    df = pd.read_csv(file_path)
                else:
                    df = pd.read_excel(file_path)

                # Establecer la conexión a la base de datos PostgreSQL
                engine = create_engine(f"postgresql://{user4}:{passw4}@localhost:5432/sistema")

                # Insertar los datos en la tabla existente (reemplaza "mi_tabla" con el nombre de tu tabla)
                df.to_sql(table_n1, con=engine, if_exists="append", index=False)

                # Mostrar una alerta de éxito
                messagebox.showinfo("Éxito", "Datos importados correctamente.")

            except Exception as e:
                # Mostrar una alerta de error
                messagebox.showerror("Error", f"Error al importar datos:\n{str(e)}")

        # Función para importar datos a PostgreSQL
        def import_data_to_postgres():
            user4= e1.get()
            passw4 = e2.get()
            table_n1='alumnos'
            try:
                # Abrir el cuadro de diálogo para seleccionar el archivo Excel o CSV
                file_path = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx;*.xls"), ("Archivos CSV", "*.csv")])

                # Leer el archivo con pandas
                if file_path.endswith(".csv"):
                    df = pd.read_csv(file_path)
                else:
                    df = pd.read_excel(file_path)

                # Establecer la conexión a la base de datos PostgreSQL
                engine = create_engine(f"postgresql://{user4}:{passw4}@localhost:5432/sistema")

                # Insertar los datos en la tabla existente (reemplaza "mi_tabla" con el nombre de tu tabla)
                df.to_sql(table_n1, con=engine, if_exists="append", index=False)

                # Mostrar una alerta de éxito
                messagebox.showinfo("Éxito", "Datos importados correctamente.")

            except Exception as e:
                # Mostrar una alerta de error
                messagebox.showerror("Error", f"Error al importar datos:\n{str(e)}")

        # Crear una ventana secundaria.
        ventana_INSERTAR = tk.Toplevel()
        #color de fondo BG
        ventana_INSERTAR.config(bg=color_defondo)
        #se le añade un titulo a la app
        ventana_INSERTAR.title("INSERTAR REGISTROS")
        ventana_INSERTAR.geometry("500x300+400+150")

        tk.Label(ventana_INSERTAR, text="Recuerde que solo se carga la tabla completa si desea ingresar un registro nuevo ", bg="#1D1212", fg="RED").place(x=10, y=5)
        tk.Label(ventana_INSERTAR, text="elimine la tabla y añada el registro en la tabla de excel ", bg="#1D1212", fg="RED").place(x=10, y=20)
        
        tk.Label(ventana_INSERTAR, text="ELija una Opcion", bg="#1D1212", fg="white").place(x=50, y=50)
        # Crear un botón dentro de la ventana secundaria
        # para cerrar la misma.
        btn_Alumnos_insert = tk.Button(
            ventana_INSERTAR,
            text="Cargar Nuevos Alumnos",
            bg="#0020BE", fg="white", relief=tk.RAISED, bd=5, 
            command=lambda: import_data_to_postgres()
        )
        btn_Conceptos_insert = tk.Button(
            ventana_INSERTAR,
            text="Cargar Nuevos Conceptos",
            bg="#0020BE", fg="white", relief=tk.RAISED, bd=5,
            command=lambda: import_data_con()
        )
        btn_cerrar_insert = tk.Button(
            ventana_INSERTAR,
            text="Cerrar ventana",
            bg="#0020BE", fg="white", relief=tk.RAISED, bd=5, 
            command=ventana_INSERTAR.destroy
        )

        btn_cerrar_insert.place(x=50, y=150, width=90, height=40)
        btn_Alumnos_insert.place(x=150, y=150, width=145, height=40)
        btn_Conceptos_insert.place(x=310, y=150, width=150, height=40)

        # Cierra la ventana principal
        ventana.withdraw()

    # Definir la función que se ejecutará cuando se haga clic en el botón
    def enviar_datos():
        #conexion a la BD
        user3 = e1.get()
        passw3 = e2.get()
        bd="sistema"
        try:
            conexion = psycopg2.connect(
                                user=user3,
                                password=passw3,
                                host='localhost',
                                port='5432',
                                database= bd
                            )
            messagebox.showinfo("CONEXION BD", "REGISTRO INGRESADO CORRECTAMENTE.")
            # Crear un cursor para ejecutar consultas SQL
            cursor = conexion.cursor()
        except:
            messagebox.askretrycancel("ERROR", "OCURRIO UN ERROR AL INGRESAR LOS DATOS, PORFAVOR CORROBORE QUE TODOS LOS DATOS SEAN CORRECTOS")
        
        FOL = folio.get("1.0", tk.END)
        MATRI = mat.get()
        NOMA = nom.get()
        GRUP = grup.get()
        LIC = lice.get()
        FE = date1
        PLA = plan.get()
        USU = user1
        ESTATUS_FOL ="ACTIVO"
        #Conceptos
        CON1 = conc1.get()
        CON2 = conc2.get()
        CON3 = conc3.get()
        CON4 = conc4.get()
        CON5 = conc5.get()
        CONCEPTOF= CON1 + "," + CON2 + "," + CON3 + "," + CON4 + "," + CON5

        TOTAL= float(total1.get("1.0", tk.END))
        #Cantidad que recibe
        CANTR = float(cant_res.get("1.0", tk.END))
        #forma de pago
        FP = formP.get()
        #Aprovacion
        APRO = aprov.get()
        #cuenta reseptora
        CUR = cuenta_res.get()
        #Observaciones
        OBSERVA = observ.get()
        # Crear un cursor para ejecutar consultas SQL
        cursor = conexion.cursor()

        try:
            # Crear un cursor para ejecutar consultas SQL
            cursor = conexion.cursor()
            # Ejecutar la consulta de inserción (adaptar según tu tabla)
            consulta = "INSERT INTO ventas (folio, plantel, matri, nombre, grupo, licenciatura, cant_recibe, forma_p, aprov, cuenta_rec, observaciones, subtotal, fecha, concep, usuario, estatus_fol) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s)"
            # Imprimir la consulta y los valores antes de ejecutarla
            cursor.execute(consulta, (FOL, PLA, MATRI, NOMA, GRUP, LIC, CANTR, FP, APRO, CUR, OBSERVA, TOTAL, FE, CONCEPTOF, USU, ESTATUS_FOL))

            # Confirmar la transacción
            conexion.commit()
            print("Datos ingresados correctamente.")
        except Exception as e:
            # Manejar cualquier error que pueda ocurrir durante la inserción
            print(f"Error al ingresar datos: {e}")
            conexion.rollback()

        # Ventana de confirmación para guardar PDF
        confirmacion_pdf()

        # Cerrar la conexión y el cursor
        cursor.close()
        conexion.close()

    def LIMPIAR_DATOS():
            numeral = 0.00
        # Limpiar campos de entrada después de la inserción
            folio.delete(1.0, tk.END)
            #folio.insert(tk.END, "")
            plan.set("")
            mat.delete(0, tk.END)
            nom.set("")
            grup.set("")
            lice.set("")
            dateUno.set("")
            dateUno.set(date1)

            conc1.delete(0, tk.END)
            conc2.delete(0, tk.END)
            conc3.delete(0, tk.END)
            conc4.delete(0, tk.END)
            conc5.delete(0, tk.END)

            impo1.delete(1.0, tk.END)
            impo2.delete(1.0, tk.END)
            impo3.delete(1.0, tk.END)
            impo4.delete(1.0, tk.END)
            impo5.delete(1.0, tk.END)

            impo1.insert(tk.END, f"{numeral:.2f}")
            impo2.insert(tk.END, f"{numeral:.2f}")
            impo3.insert(tk.END, f"{numeral:.2f}")
            impo4.insert(tk.END, f"{numeral:.2f}")
            impo5.insert(tk.END, f"{numeral:.2f}")

            combo1.delete(1.0, tk.END)
            combo2.delete(1.0, tk.END)
            combo3.delete(1.0, tk.END)
            combo4.delete(1.0, tk.END)
            combo5.delete(1.0, tk.END)

            combo1.insert(tk.END, f"{numeral:.2f}")
            combo2.insert(tk.END, f"{numeral:.2f}")
            combo3.insert(tk.END, f"{numeral:.2f}")
            combo4.insert(tk.END, f"{numeral:.2f}")
            combo5.insert(tk.END, f"{numeral:.2f}")

            rec1.delete(1.0, tk.END)
            rec2.delete(1.0, tk.END)
            rec3.delete(1.0, tk.END)
            rec4.delete(1.0, tk.END)
            rec5.delete(1.0, tk.END)

            rec1.insert(tk.END, f"{numeral:.2f}")
            rec2.insert(tk.END, f"{numeral:.2f}")
            rec3.insert(tk.END, f"{numeral:.2f}")
            rec4.insert(tk.END, f"{numeral:.2f}")
            rec5.insert(tk.END, f"{numeral:.2f}")

            subt1.delete(1.0, tk.END)
            subt2.delete(1.0, tk.END)
            subt3.delete(1.0, tk.END)
            subt4.delete(1.0, tk.END)
            subt5.delete(1.0, tk.END)

            cant_res.delete(1.0, tk.END)
            cant_res.insert(tk.END, f"{numeral:.2f}")
            cambio.delete(1.0, tk.END)
            cambio.insert(tk.END, f"{numeral:.2f}")
            total1.delete(1.0, tk.END)
            formP.delete(0,tk.END)
            aprov.delete(0,tk.END)
            cuenta_res.delete(0,tk.END)
            observ.delete(0,tk.END)

    def obtener_opciones_bd_combo():
        user3 = e1.get()
        passw3 = e2.get()
        bd="sistema"
        conexion = psycopg2.connect(
                            user=user3,
                            password=passw3,
                            host='localhost',
                            port='5432',
                            database= bd
                        )

    # Crear un cursor para ejecutar consultas SQL
        cursor = conexion.cursor()

        # Ejecutar una consulta para obtener opciones (reemplaza con tu propia consulta)
        cursor.execute("SELECT concepto FROM conceptos")
        opciones = cursor.fetchall()

        # Cerrar la conexión y el cursor
        cursor.close()
        conexion.close()

        return opciones
    
    def mostrar_opciones_combo():
        # Obtener las opciones de la base de datos
        opciones_bd_con = obtener_opciones_bd_combo()

        # Limpiar el ComboBox
        conc1['values'] = ()
        conc2['values'] = ()
        conc3['values'] = ()
        conc4['values'] = ()
        conc5['values'] = ()

        # Configurar las nuevas opciones en el ComboBox
        conc1['values'] = [opcion[0] for opcion in opciones_bd_con]
        conc2['values'] = [opcion[0] for opcion in opciones_bd_con]
        conc3['values'] = [opcion[0] for opcion in opciones_bd_con]
        conc4['values'] = [opcion[0] for opcion in opciones_bd_con]
        conc5['values'] = [opcion[0] for opcion in opciones_bd_con]
    
    def obtener_registro_completo(llave_primaria):
        user3 = e1.get()
        passw3 = e2.get()
        bd="sistema"
        conexion = psycopg2.connect(
                            user=user3,
                            password=passw3,
                            host='localhost',
                            port='5432',
                            database= bd
                        )

        cursor = conexion.cursor()

        cursor.execute("SELECT * FROM alumnos WHERE Matricula = %s", (llave_primaria,))
        registro_completo = cursor.fetchone()

        cursor.close()
        conexion.close()

        return registro_completo
    
    def obtener_registro_completo_concep(OPCC):
            user3 = e1.get()
            passw3 = e2.get()
            bd="sistema"
            conexion = psycopg2.connect(
                                user=user3,
                                password=passw3,
                                host='localhost',
                                port='5432',
                                database= bd
                            )

            cursor = conexion.cursor()
            cursor.execute("SELECT * FROM conceptos WHERE concepto = %s", (OPCC,))
            RC = cursor.fetchone()

            cursor.close()
            conexion.close()

            return RC
    
    def accion_al_seleccionar_matInicial(event):
         # Vectores para almacenar los valores
        matriculas = []
        nombres = []
        plantels = []
        oferta_ed =[]
        gra =[]
        gru =[]
        periodo =[]
        turno =[]
        fecha_in =[]
        estatus=[]
        fecha_es=[]
    # Añade más vectores según sea necesario para las demás columnas
        opcion_seleccionada = mat.get()
        # Actualizar el contenido del Label con la opción seleccionada
        #label_resultado.config(text=f"Opción seleccionada: {opcion_seleccionada}")
        if opcion_seleccionada:
            registro_completo = obtener_registro_completo(opcion_seleccionada)
            registro_formateado = ', '.join(map(str, registro_completo))
            if registro_completo:
                # Almacena cada valor del registro en vectores
                Matricula, alumno, plantel,oferta, grado, grupo1, peri, turn, date_in, sta, date_es = registro_completo
                matriculas.append(Matricula)
                nombres.append(alumno)
                plantels.append(plantel)
                oferta_ed.append(oferta) 
                gra.append(grado)
                gru.append(grupo1)
                periodo.append(peri)
                turno.append(turn)
                fecha_in.append(date_in)
                estatus.append(sta)
                fecha_es.append(''.join(map(str, date_es)))
                # Formatea la cadena antes de imprimir
                registro_formateado = ', '.join(map(str, registro_completo))
                #segun la matricula imprimimos la informacion del alumno
                plan.set(str(registro_completo[2]))
                nom.set(str(registro_completo[1]))
                grup.set(str(registro_completo[5]))
                lice.set(str(registro_completo[3]))
            else:
                print("Registro no encontrado.")

    def opc_con1(event):
         # Vectores para almacenar los valores
        cod = []
        conce = []
        cuen_con = []
        pressc =[]
        clasi =[]
    # Añade más vectores según sea necesario para las demás columnas
        opcion_seleccionada = str(conc1.get())
        # Actualizar el contenido del Label con la opción seleccionada
        #label_resultado.config(text=f"Opción seleccionada: {opcion_seleccionada}")
        if opcion_seleccionada:
            registro_completo = obtener_registro_completo_concep(opcion_seleccionada)
            registro_formateado = ', '.join(map(str, registro_completo))
            if registro_completo:
                # Almacena cada valor del registro en vectores
                Codigo, concepto, cuenta_con,precio, clasificac = registro_completo
                cod.append(Codigo)
                conce.append(concepto)
                cuen_con.append(cuenta_con)
                pressc.append(precio) 
                clasi.append(clasificac)
                # Formatea la cadena antes de imprimir
                registro_formateado = ', '.join(map(str, registro_completo))
                #segun la matricula imprimimos la informacion del alumno
                #impo1.config(text=str(registro_completo[3]))
                impo1.config(state=tk.NORMAL)
                impo1.delete(1.0, tk.END)
                impo1.insert(tk.END, float(registro_completo[3]))
            else:
                print("Registro no encontrado.")
    def opc_con2(event):
    
         # Vectores para almacenar los valores
        cod = []
        conce = []
        cuen_con = []
        pressc =[]
        clasi =[]
    # Añade más vectores según sea necesario para las demás columnas
        opcion_seleccionada = str(conc2.get())
        # Actualizar el contenido del Label con la opción seleccionada
        #label_resultado.config(text=f"Opción seleccionada: {opcion_seleccionada}")
        if opcion_seleccionada:
            registro_completo = obtener_registro_completo_concep(opcion_seleccionada)
            registro_formateado = ', '.join(map(str, registro_completo))
            if registro_completo:
                # Almacena cada valor del registro en vectores
                Codigo, concepto, cuenta_con,precio, clasificac = registro_completo
                cod.append(Codigo)
                conce.append(concepto)
                cuen_con.append(cuenta_con)
                pressc.append(precio) 
                clasi.append(clasificac)
                # Formatea la cadena antes de imprimir
                registro_formateado = ', '.join(map(str, registro_completo))
                #segun la matricula imprimimos la informacion del alumno
                #impo1.config(text=str(registro_completo[3]))
                impo2.config(state=tk.NORMAL)
                impo2.delete(1.0, tk.END)
                impo2.insert(tk.END, float(registro_completo[3]))
            else:
                print("Registro no encontrado.")
    def opc_con3(event):
         # Vectores para almacenar los valores
        cod = []
        conce = []
        cuen_con = []
        pressc =[]
        clasi =[]
    # Añade más vectores según sea necesario para las demás columnas
        opcion_seleccionada = str(conc3.get())
        # Actualizar el contenido del Label con la opción seleccionada
        #label_resultado.config(text=f"Opción seleccionada: {opcion_seleccionada}")
        if opcion_seleccionada:
            registro_completo = obtener_registro_completo_concep(opcion_seleccionada)
            registro_formateado = ', '.join(map(str, registro_completo))
            if registro_completo:
                # Almacena cada valor del registro en vectores
                Codigo, concepto, cuenta_con,precio, clasificac = registro_completo
                cod.append(Codigo)
                conce.append(concepto)
                cuen_con.append(cuenta_con)
                pressc.append(precio) 
                clasi.append(clasificac)
                # Formatea la cadena antes de imprimir
                registro_formateado = ', '.join(map(str, registro_completo))
                #segun la matricula imprimimos la informacion del alumno
                #impo1.config(text=str(registro_completo[3]))
                impo3.config(state=tk.NORMAL)
                impo3.delete(1.0, tk.END)
                impo3.insert(tk.END, float(registro_completo[3]))
            else:
                print("Registro no encontrado.")
    def opc_con4(event):
         # Vectores para almacenar los valores
        cod = []
        conce = []
        cuen_con = []
        pressc =[]
        clasi =[]
    # Añade más vectores según sea necesario para las demás columnas
        opcion_seleccionada = str(conc4.get())
        # Actualizar el contenido del Label con la opción seleccionada
        #label_resultado.config(text=f"Opción seleccionada: {opcion_seleccionada}")
        if opcion_seleccionada:
            registro_completo = obtener_registro_completo_concep(opcion_seleccionada)
            registro_formateado = ', '.join(map(str, registro_completo))
            if registro_completo:
                # Almacena cada valor del registro en vectores
                Codigo, concepto, cuenta_con,precio, clasificac = registro_completo
                cod.append(Codigo)
                conce.append(concepto)
                cuen_con.append(cuenta_con)
                pressc.append(precio) 
                clasi.append(clasificac)
                # Formatea la cadena antes de imprimir
                registro_formateado = ', '.join(map(str, registro_completo))
                #segun la matricula imprimimos la informacion del alumno
                #impo1.config(text=str(registro_completo[3]))
                impo4.config(state=tk.NORMAL)
                impo4.delete(1.0, tk.END)
                impo4.insert(tk.END, float(registro_completo[3]))
            else:
                print("Registro no encontrado.")
    def opc_con5(event):
         # Vectores para almacenar los valores
        cod = []
        conce = []
        cuen_con = []
        pressc =[]
        clasi =[]
    # Añade más vectores según sea necesario para las demás columnas
        opcion_seleccionada = str(conc5.get())
        # Actualizar el contenido del Label con la opción seleccionada
        #label_resultado.config(text=f"Opción seleccionada: {opcion_seleccionada}")
        if opcion_seleccionada:
            registro_completo = obtener_registro_completo_concep(opcion_seleccionada)
            registro_formateado = ', '.join(map(str, registro_completo))
            if registro_completo:
                # Almacena cada valor del registro en vectores
                Codigo, concepto, cuenta_con,precio, clasificac = registro_completo
                cod.append(Codigo)
                conce.append(concepto)
                cuen_con.append(cuenta_con)
                pressc.append(precio) 
                clasi.append(clasificac)
                # Formatea la cadena antes de imprimir
                registro_formateado = ', '.join(map(str, registro_completo))
                #segun la matricula imprimimos la informacion del alumno
                #impo1.config(text=str(registro_completo[3]))
                impo5.config(state=tk.NORMAL)
                impo5.delete(1.0, tk.END)
                impo5.insert(tk.END, float(registro_completo[3]))
            else:
                print("Registro no encontrado.")    
    
    def obtener_mat_bd():
        user3 = e1.get()
        passw3 = e2.get()
        bd="sistema"
        conexion = psycopg2.connect(
                            user=user3,
                            password=passw3,
                            host='localhost',
                            port='5432',
                            database= bd
                        )
        # Crear un cursor para ejecutar consultas SQL
        cursor = conexion.cursor()

        # Ejecutar una consulta para obtener opciones (reemplaza con tu propia consulta)
        cursor.execute("SELECT * FROM alumnos")
        matris = cursor.fetchall()

        # Cerrar la conexión y el cursor
        cursor.close()
        conexion.close()

        return matris

    def on_key_release_matP(event):
                    number = mostrar_mat()
                    try:
                        value = int(combobox_variable.get())
                        filtered_values = [item for item in number if str(value) in str(item)]

                    except ValueError:
                        filtered_values = []

                    mat['values'] = filtered_values

    def mostrar_mat():
        # Obtener las opciones de la base de datos
        opciones_bd_mat = obtener_mat_bd()
        # Limpiar el ComboBox
        mat['values'] = ()
        # Configurar las nuevas opciones en el ComboBox
        mat['values'] = [opcion[0] for opcion in opciones_bd_mat]
        
        number_list3 = mat['values'] = [opcion[0] for opcion in opciones_bd_mat]
        
        return number_list3
        #mat.configure(values=[opcion[0] for opcion in opciones_bd_mat])

    #FUNCION PARA ACTUALIZAR LA HORA
    def times():
        current_time=time.strftime("%H:%M:%S") 
        clock.config(text=current_time, bg=color_defondo ,fg=color_letras_fon ,font=('calibri', 30, 'bold'))
        clock.after(200,times)

    def fechaactualizacion():
        dateUno.set("")
        dateUno.set(f"{date1}")
        
    #ESTAN TODAS LAS FUNCIONES PARA LOS PRECIOS/PAGOS/CAMBIO Y OBTENER EL FOLIO
    def OPERACIONES_CONTABLES():  
        def C_O_1():
            try:
                importe1 = float(impo1.get("1.0", tk.END))
                porcentaje1 = float(combo1.get("1.0", tk.END))
                recarg = float(rec1.get("1.0", tk.END))
                
                resultado1 = importe1 - (importe1 * porcentaje1 / 100)
                resultado1 = resultado1 + recarg
                #subt1.config(text=f"{resultado:.2f}")
                subt1.config(state=tk.NORMAL)
                subt1.delete(1.0, tk.END)
                subt1.insert(tk.END, f"{resultado1:.2f}")
            except ValueError:
                print("ERROR AL INGRESAR PRECIO")
            subt1.after(1000, C_O_1)
        def C_O_2():
            try:
                importe2 = float(impo2.get("1.0", tk.END))
                porcentaje2 = float(combo2.get("1.0", tk.END))
                recarg2 = float(rec2.get("1.0", tk.END))
                
                resultado2 = importe2 - (importe2 * porcentaje2 / 100)
                resultado2 = resultado2 + recarg2

                #subt1.config(text=f"{resultado:.2f}")
                subt2.config(state=tk.NORMAL)
                subt2.delete(1.0, tk.END)
                subt2.insert(tk.END, f"{resultado2:.2f}")
            except ValueError:
                print("ERROR AL INGRESAR PRECIO")
            subt2.after(1000, C_O_2)
        def C_O_3():
            try:
                importe3 = float(impo3.get("1.0", tk.END))
                porcentaje3 = float(combo3.get("1.0", tk.END))
                recarg3 = float(rec3.get("1.0", tk.END))
                
                resultado3 = importe3 - (importe3 * porcentaje3 / 100)
                resultado3 = resultado3 + recarg3
                #subt1.config(text=f"{resultado:.2f}")
                subt3.config(state=tk.NORMAL)
                subt3.delete(1.0, tk.END)
                subt3.insert(tk.END, f"{resultado3:.2f}")
            except ValueError:
                print("ERROR AL INGRESAR PRECIO")
            subt3.after(1000, C_O_3)
        def C_O_4():
            try:
                importe4 = float(impo4.get("1.0", tk.END))
                porcentaje4 = float(combo4.get("1.0", tk.END))
                recarg4 = float(rec4.get("1.0", tk.END))
                
                resultado4 = importe4 - (importe4 * porcentaje4 / 100)
                resultado4 = resultado4 + recarg4

                #subt1.config(text=f"{resultado:.2f}")
                subt4.config(state=tk.NORMAL)
                subt4.delete(1.0, tk.END)
                subt4.insert(tk.END, f"{resultado4:.2f}")
            except ValueError:
                print("ERROR AL INGRESAR PRECIO")
            subt4.after(1000, C_O_4)
        def C_O_5():
            try:
                importe = float(impo5.get("1.0", tk.END))
                porcentaje = float(combo5.get("1.0", tk.END))
                recarg5 = float(rec5.get("1.0", tk.END))
                resultado5 = importe - (importe * porcentaje / 100)
                resultado5 = resultado5 + recarg5
                #subt1.config(text=f"{resultado:.2f}")
                subt5.config(state=tk.NORMAL)
                subt5.delete(1.0, tk.END)
                subt5.insert(tk.END, f"{resultado5:.2f}")
            except ValueError:
                print("ERROR AL INGRESAR PRECIO")
            subt5.after(1000, C_O_5)
        def TOTAL_A_APAGAR():
            try:
                importe1 = float(subt1.get("1.0", tk.END))
                importe2 = float(subt2.get("1.0", tk.END))
                importe3 = float(subt3.get("1.0", tk.END))
                importe4 = float(subt4.get("1.0", tk.END))
                importe5 = float(subt5.get("1.0", tk.END))
                
                T = importe1 + importe2 + importe3 + importe4 + importe5 

                #subt1.config(text=f"{resultado:.2f}")
                total1.config(state=tk.NORMAL)
                total1.delete(1.0, tk.END)
                total1.insert(tk.END, f"{T}")
            except ValueError:
                print("ERROR AL INGRESAR PRECIO")
            subt1.after(1000, TOTAL_A_APAGAR)
        def CAMBIO1():
            try:
                varia1 = float(total1.get("1.0", tk.END))
                varia2 = float(cant_res.get("1.0", tk.END))
                Tol = varia2 - varia1
                cambio.config(state=tk.NORMAL)
                cambio.delete(1.0, tk.END)
                cambio.insert(tk.END, f"{Tol}")
            except ValueError:
                print("ERROR AL INGRESAR PRECIO")
            cambio.after(1000, CAMBIO1)
        
        C_O_1()
        C_O_2()
        C_O_3()
        C_O_4()
        C_O_5()   
        TOTAL_A_APAGAR()
        CAMBIO1()
 
    def confirmacion_pdf():
        ventana_confirmacion = tk.Toplevel(WIND)
        ventana_confirmacion.title("Confirmación")
        # Etiqueta con el mensaje
        mensaje_label = tk.Label(ventana_confirmacion, text="¿Deseas guardar el PDF?")
        mensaje_label.grid(row=0, column=0, columnspan=2, padx=20, pady=20)

        # Botón "Sí" que realiza una acción y cierra la ventana
        boton_si = tk.Button(ventana_confirmacion, text="Sí", bg="#9BD8FF", fg="black", relief=tk.RAISED, bd=5, command=lambda: guardar_pdf(ventana_confirmacion),  width=15)
        boton_si.grid(row=1, column=0, padx=10, pady=10)

        # Botón "No" que cierra la ventana sin realizar ninguna acción
        boton_no = tk.Button(ventana_confirmacion, text="No", bg="#FF4C5C", fg="black", relief=tk.RAISED, bd=5, command=ventana_confirmacion.destroy, width=15)
        boton_no.grid(row=1, column=1, padx=10, pady=10)

        # Cambiar el tamaño de la ventana
        ventana_confirmacion.geometry('280x100')
        # Centrar la ventana en la pantalla
        ventana_confirmacion.update_idletasks()
        width = ventana_confirmacion.winfo_width()
        height = ventana_confirmacion.winfo_height()
        x = (ventana_confirmacion.winfo_screenwidth() // 2) - (width // 2)
        y = (ventana_confirmacion.winfo_screenheight() // 2) - (height // 2)
        ventana_confirmacion.geometry('{}x{}+{}+{}'.format(width, height, x, y))

    # Luego, llamar a confirmacion_pdf() cuando quieras mostrar la ventana de confirmación

    def guardar_pdf(ventana_confirmacion):
    # Obtiene los datos del formulario
        datos = {
        "Folio": folio.get("1.0", tk.END),
        "Licenciatura": lice.get(),
        "Matricula": mat.get(),
        "Nombre del Alumno": nom.get(),
        #"Usuario": user.cget("text"),
        "Grupo": grup.get(),
        "Plantel": plan.get(),
        "Fecha": dateUno.cget("text"),
        "Conceptos": [conc1.get(), conc2.get(), conc3.get(), conc4.get(), conc5.get()],
        "Importes": [impo1.get("1.0", tk.END), impo2.get("1.0", tk.END), impo3.get("1.0", tk.END), impo4.get("1.0", tk.END),
                     impo5.get("1.0", tk.END)],
        "Porcentajes": [combo1.get("1.0", tk.END), combo2.get("1.0", tk.END), combo3.get("1.0", tk.END), 
                        combo4.get("1.0", tk.END), combo5.get("1.0", tk.END)],
        "Recargos": [rec1.get("1.0", tk.END), rec2.get("1.0", tk.END), rec3.get("1.0", tk.END), rec4.get("1.0", tk.END), rec5.get("1.0", tk.END)],
        "Subtotales": [subt1.get("1.0", tk.END), subt2.get("1.0", tk.END), subt3.get("1.0", tk.END),
                       subt4.get("1.0", tk.END), subt5.get("1.0", tk.END)],
        
        "Observaciones": observ.get(),
        "Total": total1.get("1.0", tk.END)
    }
 
        # Aquí puedes agregar lógica adicional para modificar o procesar los datos antes de guardarlos en el PDF
        # Solicitar al usuario el archivo PDF existente
        pdf_existente = 'img/FORMATO CONSULTOTAL.pdf'
        # Solicitar al usuario la ubicación y el nombre del archivo PDF
        nombre_pdf = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("Archivos PDF", "*.pdf")])

        if pdf_existente and nombre_pdf:
            # Crear el archivo PDF con los datos en el PDF existente
            crear_pdf(datos, nombre_pdf, pdf_existente)
        ventana_confirmacion.destroy()

    def crear_pdf(datos, nombre_pdf, pdf_existente):

        # Abrir el archivo PDF existente
        pdf_document = fitz.open(pdf_existente)

        # Obtener la primera página
        pagina = pdf_document[0]

        # Establecer el tipo de letra y tamaño
        font_name = "helvetica"  # Puedes cambiar el nombre de la fuente según tus preferencias
        font_size = 16  # Puedes cambiar el tamaño de la fuente según tus preferencias
        #, fontname=font_name, fontsize=font_size

       # Verificar si la licenciatura es larga y reemplazar con la versión corta
        licenciatura = datos['Licenciatura']
        if licenciatura == "LICENCIATURA EN CIENCIAS POLÍTICAS Y GESTIÓN PÚBLICA":
            licenciatura1 = "LIC. CIENCIAS POLITICAS Y G.P."
        elif licenciatura == "LICENCIATURA EN DERECHO EJECUTIVO":
            licenciatura1 = "LIC. EN DERECHO EJECUTIVO "
        elif licenciatura == "INGENIERÍA EN SISTEMAS COMPUTACIONALES EJECUTIVO":
            licenciatura1 = "ING. EN SISTEMAS C. EJECUTIVO"
        elif licenciatura == "LICENCIATURA EN INGENIERÍA INDUSTRIAL":
            licenciatura1 = "LIC. EN INGENIERIA INDUSTRIAL"
        elif licenciatura == "LICENCIATURA EN INGENIERÍA INDUSTRIAL EJECUTIVO":
            licenciatura1 = "LIC. EN ING. INDUSTRIAL EJECUTIVO"
        elif licenciatura == "INGENIERÍA EN SISTEMAS COMPUTACIONALES":
            licenciatura1 = "ING. EN SISTEMAS COMPUTACIONALES"
        elif licenciatura == "LICENCIATURA EN CRIMINOLOGÍA EJECUTIVO":
            licenciatura1 = "LIC. EN CRIMINOLOGÍA EJECUTIVO"
        else:
            licenciatura1 = licenciatura  # Si no es ninguna de las anteriores, mantener la licenciatura original
        
        # Agregar texto a la página
        # Puedes ajustar las coordenadas según sea necesario
        print("Licenciatura después del if:", licenciatura1)
        #ARRIBA
        pagina.insert_text((500, 45), f" {datos['Folio']}",fontname=font_name, fontsize=font_size)
        pagina.insert_text((177, 170), f" {licenciatura1}")
        pagina.insert_text((62, 122), f" {datos['Matricula']}")
        pagina.insert_text((290, 120), f" {datos['Nombre del Alumno']}")
        #pagina.insert_text((100, 620), f"Usuario: {datos['Usuario']}")
        pagina.insert_text((45, 170), f" {datos['Grupo']}")
        #pagina.insert_text((100, 580), f"Plantel: {datos['Plantel']}")
        pagina.insert_text((452, 170), f" {datos['Fecha']}")
        pagina.insert_text((25, 355), f"Observaciones: {datos['Observaciones']}")

        #ABAJO
        pagina.insert_text((500, 430), f"{datos['Folio']}", fontname=font_name, fontsize=font_size)
        pagina.insert_text((177, 550), f" {licenciatura1}")
        pagina.insert_text((62, 505), f" {datos['Matricula']}")
        pagina.insert_text((290, 505), f" {datos['Nombre del Alumno']}")
        #pagina.insert_text((100, 620), f"Usuario: {datos['Usuario']}")
        pagina.insert_text((45, 550), f" {datos['Grupo']}")
        #pagina.insert_text((100, 580), f"Plantel: {datos['Plantel']}")
        pagina.insert_text((452, 550), f" {datos['Fecha']}")
        pagina.insert_text((25, 740), f"Observaciones: {datos['Observaciones']}")


        # Insertar los Conceptos, Importes, Porcentajes, Recargos y Subtotales (ajusta las coordenadas según sea necesario)
        y_arriba = 240
        y_abajo= 622

        conceptos = datos['Conceptos']
        importes = datos['Importes']
        desc = datos['Porcentajes']
        sub = datos['Subtotales']
        rec = datos['Recargos']

        #ARRIBA
        for i in range(5):
            if conceptos [i] == "FOTO INDIVIDUAL CON BASE":
                con = "F. INDIVIDUAL CON BASE"
            else:
                con= conceptos[i]

            if importes [i].strip() == "0":
                importes1=''
            else:
                importes1=importes[i]

            if desc [i].strip() == "0":
                desc1=''
            else:
                desc1=f"{desc[i]}%"
            
            if sub [i].strip() == "0.00":
                sub1=''
            else:
                sub1=sub[i]
            
            if rec [i].strip() == "0":
                rec1=''
            else:
                rec1=rec[i]


            pagina.insert_text((30, y_arriba), f"{con}")
            pagina.insert_text((205, y_arriba), f"{importes1}")
            pagina.insert_text((315, y_arriba), f"{desc1}")
            pagina.insert_text((415, y_arriba), f"{rec1}")
            pagina.insert_text((500, y_arriba), f"{sub1}")
            y_arriba += 20

        #ABAJO
        for i in range(5):
            if conceptos [i] == "FOTO INDIVIDUAL CON BASE":
                con = "F. INDIVIDUAL CON BASE"
            else:
                con= conceptos [i]
            
            if importes [i].strip() == "0":
                importes1=''
            else:
                importes1=importes[i]

            if importes [i].strip() == "0":
                importes1=''
            else:
                importes1=importes[i]
            
            if sub [i].strip() == "0.00":
                sub1=''
            else:
                sub1=sub[i]

            if rec [i].strip() == "0":
                rec1=''
            else:
                rec1=rec[i]

            pagina.insert_text((30, y_abajo), f"{con}")
            pagina.insert_text((205, y_abajo), f"{importes1}")
            pagina.insert_text((315, y_abajo), f"{desc1}")
            pagina.insert_text((415, y_abajo), f"{rec1}")
            pagina.insert_text((500, y_abajo), f"{sub1}")
            y_abajo += 20

        #pagina.insert_text((100, 400), f"Cantidad que Recibe: {datos['Cantidad que Recibe']}")
        #pagina.insert_text((100, 380), f"Cambio: {datos['Cambio']}")
        #pagina.insert_text((100, 360), f"Forma de Pago: {datos['Forma de Pago']}")
        #pagina.insert_text((100, 340), f"Aprobación: {datos['Aprobacion']}")
        #pagina.insert_text((100, 320), f"Cuenta Receptora: {datos['Cuenta Receptora']}")
            
        pagina.insert_text((480, 335), f"{datos['Total']}")

        pagina.insert_text((480, 718), f"{datos['Total']}")
        
        # Guardar y cerrar el PDF
        pdf_document.save(nombre_pdf)
        pdf_document.close()

        messagebox.showinfo("Éxito", f"El documento ha sido creado y guardado en:\n{nombre_pdf}")
    
    def cerrar_sesion():
        WIND.destroy()
        #ventana.withdraw()
        ventana.deiconify()

    def REIMPRIMIR_RECIBO():
        def obtener_registro_completo_fol(primary_fol):
                user3 = e1.get()
                passw3 = e2.get()
                bd="sistema"
                conexion = psycopg2.connect(
                                    user=user3,
                                    password=passw3,
                                    host='localhost',
                                    port='5432',
                                    database= bd
                                )

                cursor = conexion.cursor()

                cursor.execute("SELECT * FROM ventas WHERE folio = %s", (primary_fol,))
                registro_completo_fol = cursor.fetchone()

                cursor.close()
                conexion.close()

                return registro_completo_fol

        try:
            # Conectarse a la base de datos
            user3 = e1.get()
            passw3 = e2.get()
            bd="sistema"
            conexion = psycopg2.connect(
                                    user=user3,
                                    password=passw3,
                                    host='localhost',
                                    port='5432',
                                    database= bd
                                )
            opcion_seleccionada_PKfolio = int(folio.get("1.0", tk.END))
            cursor = conexion.cursor()

            cursor.execute("SELECT concep FROM ventas WHERE folio = %s", (opcion_seleccionada_PKfolio,))
            
            # Obtener el resultado de la consulta
            resultado = cursor.fetchone()

            # Cerrar la conexión
            cursor.close()
            conexion.close()
            
            opcion_seleccionada = int(folio.get("1.0", tk.END))
            # Actualizar el contenido del Label con la opción seleccionada
            #label_resultado.config(text=f"Opción seleccionada: {opcion_seleccionada}")
            if opcion_seleccionada:
                registro_completo = obtener_registro_completo_fol(opcion_seleccionada)
                if registro_completo:
                    #segun la FOLIO imprimimos la informacion de la venta
                    entered_text_1 = str(registro_completo[1])#Planel
                    plan.set("")
                    plan.set(f"{entered_text_1}")

                    entered_text_2 = str(registro_completo[2])#MATRICULA
                    mat.config(state=tk.NORMAL)
                    mat.delete(0, tk.END)
                    mat.insert(tk.END, f"{entered_text_2}")

                    entered_text_3 = str(registro_completo[3])#Nombre
                    nom.set("")
                    nom.set(f"{entered_text_3}")

                    entered_text_4 = str(registro_completo[4])#grupo
                    grup.set("")
                    grup.set(f"{entered_text_4}")

                    entered_text_5 = str(registro_completo[5])#LICENCIATURA
                    lice.set("")
                    lice.set(f"{entered_text_5}")

                    entered_text_6 = str(registro_completo[6])#Cantidad que recibe
                    cant_res.config(state=tk.NORMAL)
                    cant_res.delete(1.0, tk.END)
                    cant_res.insert(tk.END, f"{entered_text_6}")

                    entered_text_7 = str(registro_completo[7])#FORMA DE PAGO
                    formP.config(state=tk.NORMAL)
                    formP.delete(0, tk.END)
                    formP.insert(tk.END, f"{entered_text_7}")

                    entered_text_8 = str(registro_completo[8])#APROBACION
                    aprov.delete(0, tk.END)
                    aprov.insert(tk.END, f"{entered_text_8}")

                    entered_text_9 = str(registro_completo[9])#CUENTA RECEPTORA
                    cuenta_res.config(state=tk.NORMAL)
                    cuenta_res.delete(0, tk.END)
                    cuenta_res.insert(tk.END, f"{entered_text_9}")

                    entered_text_10 = str(registro_completo[10])#OBSERVACIONES
                    observ.delete(0, tk.END)
                    observ.insert(tk.END, f"{entered_text_10}")

                    entered_text_11 = str(registro_completo[11])#TOTAL A PAGAR
                    total1.config(state=tk.NORMAL)
                    total1.delete(1.0, tk.END)
                    total1.insert(tk.END, f"{entered_text_11}")

                    entered_text_12 = str(registro_completo[12])#FECHA
                    dateUno.set("")
                    dateUno.set(f"{entered_text_12}")

                    #CONCEPTOS
                    # Si hay datos en el resultado, procesarlos
                    if resultado:
                        # Dividir la cadena por las comas y eliminar espacios en blanco
                        conceptos = str(resultado[0]).split(',')
                        # Crear un array con los conceptos y manejar campos vacíos
                        array_conceptos = [concepto.strip() if concepto.strip() != "" else "Campo Vacío" for concepto in conceptos] # Mostrar los conceptos en la consola (puedes adaptarlo según tus necesidades)
                        conc1.config(state=tk.NORMAL)
                        conc1.delete(0, tk.END)
                        conc1.insert(tk.END, f"{array_conceptos[0]}")

                        conc2.config(state=tk.NORMAL)
                        conc2.delete(0, tk.END)
                        conc2.insert(tk.END, f"{array_conceptos[1]}")

                        conc3.config(state=tk.NORMAL)
                        conc3.delete(0, tk.END)
                        conc3.insert(tk.END, f"{array_conceptos[2]}")

                        conc4.config(state=tk.NORMAL)
                        conc4.delete(0, tk.END)
                        conc4.insert(tk.END, f"{array_conceptos[3]}")

                        conc5.config(state=tk.NORMAL)
                        conc5.delete(0, tk.END)
                        conc5.insert(tk.END, f"{array_conceptos[4]}")

                     # Vectores para almacenar los valores
                    cod = []
                    conce = []
                    cuen_con = []
                    pressc =[]
                    clasi =[]
                    opcion_seleccionada1 = str(conc1.get())
                    opcion_seleccionada2 = str(conc2.get())
                    opcion_seleccionada3 = str(conc3.get())
                    opcion_seleccionada4 = str(conc4.get())
                    opcion_seleccionada5 = str(conc5.get())
                    # Actualizar el contenido del Label con la opción seleccionada
                    #label_resultado.config(text=f"Opción seleccionada: {opcion_seleccionada}")
                    if opcion_seleccionada1:
                        registro_completo_con1 = obtener_registro_completo_concep(opcion_seleccionada1)
                        if registro_completo_con1:
                            # Almacena cada valor del registro en vectores
                            Codigo, concepto, cuenta_con,precio, clasificac = registro_completo_con1
                            cod.append(Codigo)
                            conce.append(concepto)
                            cuen_con.append(cuenta_con)
                            pressc.append(precio) 
                            clasi.append(clasificac)
                            #segun la matricula imprimimos la informacion del alumno
                            impo1.config(state=tk.NORMAL)
                            impo1.delete(1.0, tk.END)
                            impo1.insert(tk.END, float(registro_completo_con1[3]))
                        else:
                            print("Registro 1 no encontrado.")
                    if opcion_seleccionada2:
                        registro_completo_con2 = obtener_registro_completo_concep(opcion_seleccionada2)
                        if registro_completo_con2:
                            # Almacena cada valor del registro en vectores
                            Codigo, concepto, cuenta_con,precio, clasificac = registro_completo_con2
                            cod.append(Codigo)
                            conce.append(concepto)
                            cuen_con.append(cuenta_con)
                            pressc.append(precio) 
                            clasi.append(clasificac)
                            #segun la matricula imprimimos la informacion del alumno
                            impo2.config(state=tk.NORMAL)
                            impo2.delete(1.0, tk.END)
                            impo2.insert(tk.END, float(registro_completo_con2[3]))
                        else:
                            print("Registro 2 no encontrado.")
                    if opcion_seleccionada3:
                        registro_completo_con3 = obtener_registro_completo_concep(opcion_seleccionada3)
                        if registro_completo_con3:
                            # Almacena cada valor del registro en vectores
                            Codigo, concepto, cuenta_con,precio, clasificac = registro_completo_con3
                            cod.append(Codigo)
                            conce.append(concepto)
                            cuen_con.append(cuenta_con)
                            pressc.append(precio) 
                            clasi.append(clasificac)
                            #segun la matricula imprimimos la informacion del alumno
                            impo3.config(state=tk.NORMAL)
                            impo3.delete(1.0, tk.END)
                            impo3.insert(tk.END, float(registro_completo_con3[3]))
                        else:
                            print("Registro 3 no encontrado.")
                    if opcion_seleccionada4:
                        registro_completo_con4 = obtener_registro_completo_concep(opcion_seleccionada4)
                        if registro_completo_con4:
                            # Almacena cada valor del registro en vectores
                            Codigo, concepto, cuenta_con,precio, clasificac = registro_completo_con4
                            cod.append(Codigo)
                            conce.append(concepto)
                            cuen_con.append(cuenta_con)
                            pressc.append(precio) 
                            clasi.append(clasificac)
                            #segun la matricula imprimimos la informacion del alumno
                            impo4.config(state=tk.NORMAL)
                            impo4.delete(1.0, tk.END)
                            impo4.insert(tk.END, float(registro_completo_con4[3]))
                        else:
                            print("Registro 4 no encontrado.")
                    if opcion_seleccionada5:
                        registro_completo_con5 = obtener_registro_completo_concep(opcion_seleccionada5)
                        if registro_completo_con5:
                            # Almacena cada valor del registro en vectores
                            Codigo, concepto, cuenta_con,precio, clasificac = registro_completo_con5
                            cod.append(Codigo)
                            conce.append(concepto)
                            cuen_con.append(cuenta_con)
                            pressc.append(precio) 
                            clasi.append(clasificac)
                            #segun la matricula imprimimos la informacion del alumno
                            impo5.config(state=tk.NORMAL)
                            impo5.delete(1.0, tk.END)
                            impo5.insert(tk.END, float(registro_completo_con5[3]))
                        else:
                            print("Registro 5 no encontrado.")
                
                else:
                    print("Registro no encontrado.")
                
                entered_text_14 = str(registro_completo[14])#usuario
                nom_usu1.set("")
                nom_usu1.set(f"{entered_text_14}")
            
        except Exception as e:
            print(f"Error: {e}")
    
    #VENTANA PRINCIPAL
    WIND = tk.Toplevel(ventana)
    #ABRIR EN PANTALLA EN GRANDE
    WIND.state("zoomed")
    # Configurar el tamaño de la ventana
    width1, height1 = 1300, 768
    WIND.geometry(f"{width1}x{height1}+650+150")

    WIND.config(bg=color_defondo)
    #se le añade un titulo a la app
    WIND.title("SISTEMA DE COBROS")
    
    #LOGO
    WIND.iconbitmap("img/univer_lg.ico")
    #la metemos dentro de un label para poder mostrarla
    eti= Label(WIND, image = img1, bg=color_defondo)
    #eti = tk.Label(WIND, image=imagen_sub)
    eti.place(x=20, y=10)

    #Formulario Crear los widgets del formulario
    paqtg1= tk.Label(WIND, text="PAQUETES DE GRADUACION", font="Candara 24 bold", bg=color_defondo, fg=color_letras_fon, width=0, height=3)
    paqtg1.place(x=850, y=-30)

    #HORA EN TIEMPO REAL
    clock=Label(WIND,font=("Candara",50,"bold"))
    clock.place(x=520, y=30)
    times()

    #USUARIO
    user1 = e1.get()
    nom_usu1 = tk.StringVar()
    nom_usu1 = user1
    enom= e0.get()
    nom_usu = tk.Label(WIND, text="Bienvenido: " + user1, bg=color_defondo, bd=3, fg=color_letras_fon, font="Candara 14 bold",textvariable=nom_usu1)
    nom_usu.place(x=850, y=50, width=400)
        
    #FOLIO
    tk.Label(WIND, text="FOLIO:", bg="#565151", fg="white", relief=tk.GROOVE, bd=3, font="Candara 16 bold").place(x=650, y=140, width=70, height=30)
    folio = tk.Text(WIND, height=1, width=10, wrap=tk.NONE, relief=tk.GROOVE, bd=4, padx=2, pady=2, font="Candara 14 bold")
    folio.place(x=725, y=140, width=100, height=30)
    folio.insert(tk.END, "")

    # Botón para REIMPRIMIR
    tk.Label(WIND, text="REIMPRIMIR RECIBO",  bg="#565151", relief=tk.GROOVE, bd=3, fg="white", font="Candara 10 bold").place(x=800, y=106, width=128, height=20)
    guardar_pdf_button = tk.Button(WIND, text="", image=imgLUPA1, compound=tk.CENTER, bg=color_defondo, fg="white", relief=tk.RAISED, bd=5, command=REIMPRIMIR_RECIBO)
    guardar_pdf_button.place(x=828, y=132, width=35, height=40)
    # Botón para guardar datos en el PDF
    tk.Label(WIND, text="IMPRIMIR",  bg="#565151", relief=tk.GROOVE, bd=3, fg="white", font="Candara 10 bold").place(x=875, y=245, width=80, height=20)
    guardar_pdf_button = tk.Button(WIND, text="", image=imgREIMPRI1, compound=tk.CENTER, bg=color_defondo, fg="white", relief=tk.RAISED, bd=5, command=guardar_pdf)
    guardar_pdf_button.place(x=890, y=185, width=50, height=60)
    #BOTON PARA LIMPIAR FORMULARIO
    tk.Label(WIND, text="LIMPIAR",  bg="#565151", relief=tk.GROOVE, bd=3, fg="white", font="Candara 10 bold").place(x=970, y=245, width=60, height=20)
    BTN_LIMPIAR = tk.Button(WIND, text="", image=imgg1, compound=tk.CENTER, bg=color_defondo, fg="white", relief=tk.RAISED, bd=5, command=LIMPIAR_DATOS)
    BTN_LIMPIAR.place(x=975, y=185, width=50, height=60)

    #MATRICULA
    opciones_bd_mat = obtener_mat_bd()
    tk.Label(WIND,text="MATRICULA:", bg="#565151", fg="white", relief=tk.GROOVE, bd=3, font="Candara 12 bold").place(x=50, y=180, width=100, height=25)
    combobox_variable = tk.StringVar()
    mat = ttk.Combobox(WIND, width=30, font="Candara 12 bold", textvariable=combobox_variable)
    mat.place(x=160, y=180, width=110, height=25)
    # Vincular la función al evento <<ComboboxSelected>>
    mat.bind("<<ComboboxSelected>>", accion_al_seleccionar_matInicial)
    mat.bind('<KeyRelease>', on_key_release_matP)
    
    #NOMBRE DEL ALUMNO
    tk.Label(WIND,text="NOMBRE DEL ALUMNO:", bg="#565151", fg="white", relief=tk.GROOVE, bd=3, font="Candara 12 bold").place(x=280, y=180, width=185, height=25)
    nom = tk.StringVar()
    nom1 = tk.Label(WIND,relief=tk.SUNKEN, bd=3, width=30, bg="white", font="Candara 11 bold", textvariable=nom)
    nom1.place(x=475, y=180, width=350, height=25)

    #LICENCIATURA
    tk.Label(WIND, text="LICENCIATURA:", bg="#565151", fg="white", relief=tk.GROOVE, bd=3, font="Candara 12 bold").place(x=50, y=220, width=125, height=25)
    lice = tk.StringVar()
    lice1 = tk.Label(WIND,relief=tk.SUNKEN, bd=3, width=42, bg="white", font="Candara 11 bold", textvariable=lice)
    lice1.place(x=185, y=220, width=395, height=25)

    #GRUPO
    tk.Label(WIND, text="GRUPO:", bg="#565151", fg="white", relief=tk.GROOVE, bd=3, font="Candara 12 bold").place(x=590, y=220, width=70, height=25)
    grup = tk.StringVar()
    grup1 = tk.Label(WIND,relief=tk.SUNKEN, bd=3, width=18, bg="white", font="Candara 11 bold", textvariable=grup)
    grup1.place(x=675, y=220, width=150, height=25)

    #PLANTEL
    tk.Label(WIND, text="PLANTEL:", bg="#565151",fg="white", relief=tk.GROOVE, bd=3, font="Candara 12 bold" ).place(x=50, y=260, width=80, height=25)
    plan = tk.StringVar()
    plan1 = tk.Label(WIND,relief=tk.SUNKEN, bd=3, width=19, bg="white", font="Candara 11 bold", textvariable=plan)
    plan1.place(x=140, y=260, width=160, height=25)

    #FECHA
    tk.Label(WIND, text="FECHA:", bg="#565151",fg="white", relief=tk.GROOVE, bd=3, font="Candara 12 bold").place(x=320, y=260, width=70, height=25)
    dateUno = tk.StringVar()
    date1 = date.today()
    dateDos= tk.Label(WIND, text=date.today(), bg=color_defondo, bd=3, fg=color_letras_fon, font="Candara 14 bold", textvariable=dateUno)
    dateDos.place(x=405, y=260, width=90, height=25)
    

    tk.Label(WIND, text="CONCEPTOS", bg="#565151", relief=tk.GROOVE, bd=3, fg="white", font="Candara 12 bold").place(x=50, y=300, width=200, height=25)
    #INICIAN LOS CONCEPTOS
    conc1 = ttk.Combobox(WIND, width=27, font="Candara 11 bold")
    conc1.place(x=50, y=340, width=200, height=25)
    # Vincular la función al evento <<ComboboxSelected>>
    conc1.bind("<<ComboboxSelected>>", opc_con1)
    conc2 = ttk.Combobox(WIND, width=27, font="Candara 11 bold")
    conc2.place(x=50, y=370, width=200, height=25)
    # Vincular la función al evento <<ComboboxSelected>>
    conc2.bind("<<ComboboxSelected>>", opc_con2)

    conc3 = ttk.Combobox(WIND, width=27, font="Candara 11 bold")
    conc3.place(x=50, y=400, width=200, height=25)
    # Vincular la función al evento <<ComboboxSelected>>
    conc3.bind("<<ComboboxSelected>>", opc_con3)

    conc4 = ttk.Combobox(WIND, width=27, font="Candara 11 bold")
    conc4.place(x=50, y=430, width=200, height=25)
    # Vincular la función al evento <<ComboboxSelected>>
    conc4.bind("<<ComboboxSelected>>", opc_con4)

    conc5 = ttk.Combobox(WIND, width=27, font="Candara 11 bold")
    conc5.place(x=50, y=460, width=200, height=25)
    # Vincular la función al evento <<ComboboxSelected>>
    conc5.bind("<<ComboboxSelected>>", opc_con5)

    #IMPORTE
    tk.Label(WIND, text="IMPORTE", bg="#565151", relief=tk.GROOVE, bd=3, fg="white", font="Candara 12 bold").place(x=270, y=300, width=79, height=25)
    impo1 = tk.Text(WIND, height=1, width=10, wrap=tk.NONE, relief=tk.SUNKEN, bd=3, padx=2, pady=2, font="Candara 11 bold")
    impo1.place(x=270, y=340, width=79, height=25)
    impo1.insert(tk.END, "0")
    
    impo2 = tk.Text(WIND, height=1, width=10, wrap=tk.NONE,  relief=tk.GROOVE, bd=3, padx=2, pady=2, font="Candara 11 bold")
    impo2.place(x=270, y=370, width=79, height=25)
    impo2.insert(tk.END, "0")

    impo3 = tk.Text(WIND, height=1, width=10, wrap=tk.NONE, relief=tk.GROOVE, bd=3, padx=2, pady=2, font="Candara 11 bold")
    impo3.place(x=270, y=400, width=79, height=25)
    impo3.insert(tk.END, "0")

    impo4 = tk.Text(WIND, height=1, width=10, wrap=tk.NONE, relief=tk.GROOVE, bd=3, padx=2, pady=2, font="Candara 11 bold")
    impo4.place(x=270, y=430, width=79, height=25)
    impo4.insert(tk.END, "0")

    impo5 = tk.Text(WIND, height=1, width=10, wrap=tk.NONE, relief=tk.GROOVE, bd=3, padx=2, pady=2, font="Candara 11 bold")
    impo5.place(x=270, y=460, width=79, height=25)
    impo5.insert(tk.END, "0")

    #%
    tk.Label(WIND, text="PORCENTAJE(%)",  bg="#565151", relief=tk.GROOVE, bd=3, fg="white", font="Candara 12 bold").place(x=370, y=300, width=115, height=25)
    combo1 = tk.Text(WIND, height=1, width=10, wrap=tk.NONE, relief=tk.GROOVE, bd=3, padx=2, pady=2, font="Candara 11 bold")
    combo1.place(x=370, y=340, width=115, height=25)
    combo1.insert(tk.END, "0")

    combo2 = tk.Text(WIND, height=1, width=10, wrap=tk.NONE, relief=tk.GROOVE, bd=3, padx=2, pady=2, font="Candara 11 bold")
    combo2.place(x=370, y=370, width=115, height=25)
    combo2.insert(tk.END, "0")

    combo3 = tk.Text(WIND, height=1, width=10, wrap=tk.NONE, relief=tk.GROOVE, bd=3, padx=2, pady=2, font="Candara 11 bold")
    combo3.place(x=370, y=400, width=115, height=25)
    combo3.insert(tk.END, "0")

    combo4 = tk.Text(WIND, height=1, width=10, wrap=tk.NONE, relief=tk.GROOVE, bd=3, padx=2, pady=2, font="Candara 11 bold")
    combo4.place(x=370, y=430, width=115, height=25)
    combo4.insert(tk.END, "0")

    combo5 = tk.Text(WIND, height=1, width=10, wrap=tk.NONE, relief=tk.GROOVE, bd=3, padx=2, pady=2, font="Candara 11 bold")
    combo5.place(x=370, y=460, width=115, height=25)
    combo5.insert(tk.END, "0")

    #RECARGOS
    tk.Label(WIND, text="RECARGOS($)", bg="#565151", relief=tk.GROOVE, bd=3, fg="white", font="Candara 12 bold").place(x=498, y=300, width=110, height=25)

    rec1 = tk.Text(WIND, height=1, width=10, wrap=tk.NONE, relief=tk.GROOVE, bd=3, padx=2, pady=2, font="Candara 11 bold")
    rec1.place(x=498, y=340, width=110, height=25)
    rec1.insert(tk.END, "0")

    rec2 = tk.Text(WIND, height=1, width=10, wrap=tk.NONE, relief=tk.GROOVE, bd=3, padx=2, pady=2, font="Candara 11 bold")
    rec2.place(x=498, y=370, width=110, height=25)
    rec2.insert(tk.END, "0")

    rec3 = tk.Text(WIND, height=1, width=10, wrap=tk.NONE, relief=tk.GROOVE, bd=3, padx=2, pady=2, font="Candara 11 bold")
    rec3.place(x=498, y=400, width=110, height=25)
    rec3.insert(tk.END, "0")

    rec4 = tk.Text(WIND, height=1, width=10, wrap=tk.NONE, relief=tk.GROOVE, bd=3, padx=2, pady=2, font="Candara 11 bold")
    rec4.place(x=498, y=430, width=110, height=25)
    rec4.insert(tk.END, "0")

    rec5 = tk.Text(WIND, height=1, width=10, wrap=tk.NONE, relief=tk.GROOVE, bd=3, padx=2, pady=2, font="Candara 11 bold")
    rec5.place(x=498, y=460, width=110, height=25)
    rec5.insert(tk.END, "0")

    #SUBTOTAL
    tk.Label(WIND, text="SUBTOTAL", bg="#565151", relief=tk.GROOVE, bd=3, fg="white", font="Candara 12 bold").place(x=620, y=300, width=90, height=25)
    #subt1 = tk.Label(WIND, width=10,relief=tk.SUNKEN, bd=3, bg="white", font="Candara 11 bold")
    subt1 = tk.Text(WIND, height=1, width=10, wrap=tk.NONE, relief=tk.GROOVE, bd=3, padx=2, pady=2, font="Candara 11 bold")
    subt1.place(x=620, y=340,width=90, height=25)
    subt1.insert(tk.END, "0")

    subt2 = tk.Text(WIND, height=1, width=10, wrap=tk.NONE, relief=tk.GROOVE, bd=3, padx=2, pady=2, font="Candara 11 bold")
    subt2.place(x=620, y=370, width=90, height=25)
    subt2.insert(tk.END, "0")

    subt3 = tk.Text(WIND, height=1, width=10, wrap=tk.NONE, relief=tk.GROOVE, bd=3, padx=2, pady=2, font="Candara 11 bold")
    subt3.place(x=620, y=400, width=90, height=25)
    subt3.insert(tk.END, "0")

    subt4 = tk.Text(WIND, height=1, width=10, wrap=tk.NONE, relief=tk.GROOVE, bd=3, padx=2, pady=2, font="Candara 11 bold")
    subt4.place(x=620, y=430, width=90, height=25)
    subt4.insert(tk.END, "0")

    subt5 = tk.Text(WIND, height=1, width=10, wrap=tk.NONE, relief=tk.GROOVE, bd=3, padx=2, pady=2, font="Candara 11 bold")
    subt5.place(x=620, y=460, width=90, height=25)
    subt5.insert(tk.END, "0")

    #TOTAL A PAGAR
    tk.Label(WIND, text="TOTAL:", bg="#565151", relief=tk.GROOVE, bd=3, fg="white", font="Candara 12 bold").place(x=720, y=300, width=165, height=25)
    total1 =  tk.Text(WIND, height=1, width=10, wrap=tk.NONE, relief=tk.GROOVE, bd=3, padx=2, pady=2, font="Candara 11 bold")
    total1.place(x=720, y=330, width=165, height=25)
    total1.insert(tk.END, "0")

    #CANTIDAD QUE RECIBE
    tk.Label(WIND, text="CANTIDAD QUE RECIBE", bg="#565151", relief=tk.GROOVE, bd=3, fg="white", font="Candara 12 bold").place(x=720, y=365, width=165, height=25)
    cant_res = tk.Text(WIND, height=1, width=10, wrap=tk.NONE,  relief=tk.GROOVE, bd=3, padx=2, pady=2, font="Candara 11 bold")
    cant_res.place(x=720, y=395, width=165, height=25)
    cant_res.insert(tk.END, "0")

    #CAMBIO
    tk.Label(WIND, text="CAMBIO", bg="#565151", relief=tk.GROOVE, bd=3, fg="white", font="Candara 12 bold").place(x=720, y=430, width=165, height=25)
    cambio = tk.Text(WIND, height=1, width=10, wrap=tk.NONE, relief=tk.SUNKEN, bd=3, padx=2, pady=2, font="Candara 11 bold")
    cambio.place(x=720, y=460, width=165, height=25)
    cambio.insert(tk.END, "0")

    #FORMA DE PAGO
    tk.Label(WIND, text="FORMA DE PAGO", bg="#565151", relief=tk.GROOVE, bd=3, fg="white", font="Candara 12 bold").place(x=895, y=300, width=152, height=25)
    formP = ttk.Combobox( WIND, state="normal",font="Candara 11 bold", values=["Efectivo", "Tarjeta de Credito", "Tarjeta de Debito", "Cheque", "Deposito Bancario"])
    formP.place(x=895, y=330, width=152, height=25)

    #APROVACION
    aprov = tk.Label(WIND, text="APROBACIÓN", bg="#565151", relief=tk.GROOVE, bd=3, fg="white", font="Candara 12 bold").place(x=895, y=365, width=152, height=25)
    aprov = tk.Entry(WIND, width=25,relief=tk.SUNKEN, bd=4, font="Candara 11 bold")
    aprov.place(x=895, y=395, width=152, height=25)

    #CUENTA RECEPTORA
    cuenta_res = tk.Label(WIND, text="CUENTA RECEPTORA", bg="#565151", relief=tk.GROOVE, bd=3, fg="white", font="Candara 12 bold").place(x=895, y=430, width=152, height=25)
    cuenta_res = ttk.Combobox(WIND,state="readonly",values=["Santander", "No aplica"], font="Candara 11 bold")
    cuenta_res.place(x=895, y=460, width=152, height=25)

    #OBSERVACIONES
    tk.Label(WIND, text="OBSERVACIONES:",  bg="#565151", relief=tk.GROOVE, bd=3, fg="white", font="Candara 12 bold").place(x=50, y=505, width=137, height=25)
    observ = tk.Entry(WIND, width=30, relief=tk.GROOVE, bd=4, font="Candara 11 bold")
    observ.place(x=200, y=505, width=690, height=25)

    #BOTONES
    boton_enviar = tk.Button(WIND, text="REGISTRAR", bg="#11BE00", fg="white", relief=tk.RAISED, bd=5,font="Candara 15 bold", command=enviar_datos)
    boton_enviar.place(x=920, y=495, width=115, height=35)

    btn_Consultar = tk.Button(WIND, text="CONSULTAR", bg="#00BBBE", fg="white", relief=tk.RAISED, bd=5,font="Candara 15 bold", command=Consultar)
    btn_Consultar.place(x=210, y=600, width=115, height=35)

    btn_Insert = tk.Button(WIND, text="CARGAR", bg="#00BE7C", fg="white", relief=tk.RAISED, bd=5,font="Candara 15 bold", command=Insertar)
    btn_Insert.place(x=330, y=600, width=115, height=35)

    btn_Actualizar = tk.Button(WIND, text="ACTUALIZAR", bg="#E37F00", fg="white", relief=tk.RAISED, bd=5,font="Candara 15 bold", command=actualizar)
    btn_Actualizar.place(x=460, y=600, width=115, height=35)

    btn_eliminar = tk.Button(WIND, text="ELIMINAR", bg="#CA0000", fg="white", relief=tk.RAISED, bd=5,font="Candara 15 bold", command=eliminar)
    btn_eliminar.place(x=580, y=600, width=115, height=35)

    #CERRAR SESION
    tk.Label(WIND, text="CERRAR SESIÓN",  bg="#565151", relief=tk.GROOVE, bd=3, fg="white", font="Candara 10 bold").place(x=1250, y=600, width=110, height=20)
    btn_cerrar_sesion = tk.Button(WIND, text="", image=imgSES1, compound=tk.CENTER, bg="#1D1212", fg="white", relief=tk.RAISED, bd=5, command=cerrar_sesion)
    btn_cerrar_sesion.place(x=1300, y=630, width=50, height=50)
    # Llamar a las funciones al inicio
    OPERACIONES_CONTABLES()
    mostrar_opciones_combo()
    mostrar_mat()
    fechaactualizacion()
    # Cierra la ventana principal
    ventana.withdraw()

#LOGIN
def Evalue():
    user3 = e1.get()
    passw3 = e2.get()
    nombre = e0.get()
    bd="sistema"
    # Verifica si los campos están vacíos
    if not nombre or not user3 or not passw3:
        messagebox.showinfo("Campo Vacío", "Por favor, llene todos los campos.")
        return
    try:
        conexion = psycopg2.connect(
                            user=user3,
                            password=passw3,
                            host='localhost',
                            port='5432',
                            database= bd
                        )
        messagebox.showinfo("INICIO DE SESION", "DATOS CORRECTOS")
        INDEX()
    except:
        messagebox.askretrycancel("ERROR", "Inicio de Sesion fallido, Intentelo de nuevo.")
    # Crear un cursor para ejecutar consultas SQL
    cursor = conexion.cursor()
    # Cerrar la conexión y el cursor
    cursor.close()
    conexion.close()

def toggle_visibility():
    if e2.cget("show") == "*":
        eye_button.config(image=open_eye)
        e2.configure(show="")
    else:
        eye_button.config(image=close_eye)
        e2.configure(show="*")

def cerrar_ventana():
    ventana.destroy()
ventana = tk.Tk()
#tamaño de la ventana
ventana.geometry("520x500")
#color de fondo BG "#D3D3D3" #0D0D0D
color_defondo = "#D3D3D3"
color_letras_fon = "#4169e1"
#color de fondo BG
ventana.config(bg="#1D1212")
#se le añade un titulo a la app
ventana.title("SISTEMA DE COBROS")
#Evita que se pueda redimensionar
ventana.resizable(False,False)
# Quita los botones de minimizar y maximizar
#ventana.overrideredirect(True)

# Centrar la ventana en la pantalla
ventana.update_idletasks()
width = ventana.winfo_width()
height = ventana.winfo_height()
x = (ventana.winfo_screenwidth() // 2) - (width // 2)
y = (ventana.winfo_screenheight() // 2) - (height // 2)
ventana.geometry(f'{width}x{height}+{x}+{y}')
#LOGO
#ventana.iconbitmap("img/LG.ico")

#Frame
frame= CTkFrame(ventana, fg_color="#1D1212")
frame.grid(column=0, row=0, sticky='nsew', padx=50,pady=50)

# Crear una instancia de CTkFont
font = ctk.CTkFont(family="sans rerif", size=12)

# Agrega un campo de entrada para el nombre
label_nombre = ctk.CTkLabel(frame, text="NOMBRE", font=("Candara", 20, "bold"), bg_color="#1D1212")
label_nombre.grid(row=1, column=0, padx=4, pady=10)

e0 = ctk.CTkEntry(frame, font=font, placeholder_text='INGRESA TU NOMBRE', 
                  border_color="#4169e1", fg_color="#1D1212", width=220, height=40)
e0.grid(row=1, column=1, padx=4, pady=10)

#USUARIO
usuario= ctk.CTkLabel(frame, text="USER", font=("Candara", 20, "bold"), bg_color="#1D1212")
usuario.grid(row=2, column=0, padx=4, pady=10)

#USUARIO ENTRADA
e1 = CTkEntry(frame, font=font, placeholder_text='USUARIO',
              border_color="#4169e1", fg_color="#1D1212", width=220, height=40)
e1.grid(row=2, column=1, padx=4, pady=10)

#CONTRASENA
contra = CTkLabel(frame, text="PASSWORD", font=("Candara", 20, "bold"), bg_color="#1D1212", height=10, width=100)
contra.grid(row=3, column=0, padx=4, pady=10)

#CONTRASENA  ENTRADA 
e2 = CTkEntry(frame, font=font, show='*', placeholder_text='CONTRASENA',
              border_color="#4169e1", fg_color="#1D1212", width=220, height=40)
e2.grid(row=3, column=1, padx=4, pady=4)

# IMAGEN adding image (remember image should be PNG and not JPG)
img = PhotoImage(file='img/LOGO_TRANSPARENTE.png')
img1 = img.subsample(2, 2)
Label(frame,image=img1, bg="#1D1212").grid(columnspan=2, row=0)

ventana.columnconfigure(0, weight=1)
ventana.rowconfigure(0,weight=1)


# Crea un objeto CTkFont con las características deseadas
font = ctk.CTkFont(family="calibri", size=16, weight="bold")

# Crea un botón personalizado
b1 = ctk.CTkButton(frame,
                font=font,
                text="LOGIN",
                text_color="white",
                border_color="#4169e1", 
                fg_color="#1D1212", 
                hover_color="#4169e1",
                corner_radius=12,
                border_width=2,
                command=Evalue
)

# Agrega el botón a la ventana
b1.grid(row=4, column=0, columnspan=2, padx=4, pady=20)

# Asocia la tecla Enter a la función Evalue
e1.bind("<Return>", lambda event: Evalue())
e2.bind("<Return>", lambda event: Evalue())

# Ojo para mostrar/ocultar contraseña
eye_button = Button(frame, command=toggle_visibility)
open_eye = PhotoImage(file='img/visible.png')  # Ajusta la ruta y el nombre de la imagen
close_eye = PhotoImage(file='img/ojo.png')  # Ajusta la ruta y el nombre de la imagen
# Cargar la imagen (reemplaza 'ruta_de_tu_imagen.png' con la ruta de tu propia imagen)
imgg = PhotoImage(file='img/limpiar.png')
imgg1 = imgg.subsample(12, 12)

imgExit = PhotoImage(file='img/salir_1.png')
imgExit1 = imgExit.subsample(12, 12)

imgSES = PhotoImage(file='img/cerrar_ses.png')
imgSES1 = imgSES.subsample(8, 8)

imgPDF = PhotoImage(file='img/pdf_icon.png')
imgPDF1 = imgPDF.subsample(5, 5)

imgREIMPRI = PhotoImage(file='img/reimprimir_ico.png')
imgREIMPRI1 = imgREIMPRI.subsample(11, 11)


imgLUPA = PhotoImage(file='img/buscar.png')#<a href="https://www.flaticon.es/iconos-gratis/buscar" title="buscar iconos">Buscar iconos creados por DinosoftLabs - Flaticon</a>
imgLUPA1 = imgLUPA.subsample(2, 2)


# Agrega un botón de cierre personalizado
boton_cerrar = tk.Button(ventana, text="", image=imgExit1, compound=tk.CENTER, bg="#1D1212", fg="white", relief=tk.RAISED, bd=5,command=cerrar_ventana)
boton_cerrar.place(x=10, y=400)

eye_button.config(image=close_eye, bd=0, bg="white")
eye_button.grid(row=3, column=1, padx=(280, 10)) # Ajusta las coordenadas según sea necesario
ventana.grid_columnconfigure(2, weight=1)  # Ajusta el peso de la columna para centrar

#moverse con las flechas para el label y boton 
def abajo(event):
    event.widget.tk_focusNext().focus()
    limite = event.widget.tk_focusNext()
    if isinstance(limite, tk.Entry):
        limite.focus()
    return "break"  # Evita que el evento se propague

def arriba(event):
    limite2 = event.widget.tk_focusPrev()
    if isinstance(limite2, tk.Entry):
        limite2.focus()
    return "break"
# Asocia el evento <Down> a la función para cada Entry
e0.bind("<Down>", abajo)
e0.bind("<Up>", arriba)
e1.bind("<Down>", abajo)
e1.bind("<Up>", arriba)
e2.bind("<Down>", abajo)
e2.bind("<Up>", arriba)
# Boton a label
b1.bind("<Up>", lambda event: e2.focus())
ventana.mainloop()
